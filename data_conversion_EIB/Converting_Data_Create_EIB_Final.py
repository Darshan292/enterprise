import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re
import numpy as np


def convert_and_format_date(input_data, column):
    # Convert column to datetime format
    input_data[column] = pd.to_datetime(input_data[column], errors='coerce')
    # Format column as "YYYY-MMM-DD"
    input_data[column] = input_data[column].dt.strftime('%Y-%m-%d')


def remove_non_numeric_char(input_data, column):
    input_data[column] = input_data[column].apply(
        lambda x: re.sub(r'[^0-9]','', str(x)) if not pd.isna(x) and str(
            x).strip() != '' else np.nan)


def convert_column_to_row(input_data, id_vars, columns_to_melt, var_name, value_name):
    
    # Check if all specified columns are null or blank for each row
    is_null_or_blank = input_data[columns_to_melt].isnull() | input_data[columns_to_melt].eq('')
    all_null_or_blank = is_null_or_blank.all(axis=1)
    
    # Replace the first column with 'removeString' if all specified columns are null or blank for a row
    input_data.loc[all_null_or_blank, columns_to_melt[0]] = 'removeString15454'

    # Select the columns to melt
    melted_data = pd.melt(input_data, id_vars=id_vars, value_vars=columns_to_melt, var_name=var_name,
                          value_name=value_name)

    # Replace blank values with NaN
    # melted_data[value_name] = melted_data[value_name].replace('', np.nan)

    # Remove rows with NaN or blank values in 'Value' column
    melted_data = melted_data.dropna(subset=[value_name])

    # Replace 'removeString' with an empty string in the 'value_name' column
    melted_data[value_name] = melted_data[value_name].replace('removeString15454', '')

    # Sort the melted data by 'Legacy Worker ID' and 'Value'
    melted_data = melted_data.sort_values(by=id_vars).reset_index(drop=True)

    return melted_data

def increment_row_id_for_same_sort_col(df, groupby_col, sort_col,row_id_col):
    """
    Generate 'row id' that increments only when the 'sort_col' value same between
    the current row and the previous row within each group of 'groupby_col'.

    Args:
    - df: DataFrame containing the data
    - groupby_col: Column to group by ('Legacy Worker ID' in this case)
    - sort_col: Column to use for sorting ('test' in this case)

    Returns:
    - DataFrame with 'row id' column added
    """
    df[row_id_col] = (df.groupby(groupby_col)[sort_col]
                    .apply(lambda x: (x == x.shift()).cumsum()+1)
                    .reset_index(drop=True))

    # Blank out the 'row_id_col' if df[sort_col] is NaN or blank
    df.loc[df[sort_col].isnull() | (df[sort_col] == ''), row_id_col] = None

    return df

def generate_row_id(df, groupby_col, sort_col,row_id_col):
    """
    Generate 'row id' that increments only when the 'sort_col' value changes between
    the current row and the previous row within each group of 'groupby_col'.

    Args:
    - df: DataFrame containing the data
    - groupby_col: Column to group by ('Legacy Worker ID' in this case)
    - sort_col: Column to use for sorting ('test' in this case)

    Returns:
    - DataFrame with 'row id' column added
    """
    df[row_id_col] = (df.groupby(groupby_col)[sort_col]
                    .apply(lambda x: (x != x.shift()).cumsum())
                    .reset_index(drop=True))

    # Blank out the 'row_id_col' if df[sort_col] is NaN or blank
    df.loc[df[sort_col].isnull() | (df[sort_col] == ''), row_id_col] = None

    return df

def remove_duplicate_details(df, worker_id_col, phone_id_col, phone_details_cols):
    """
    Removes duplicate phone number details based on the phone ID for each worker.

    Args:
    - df (DataFrame): The input DataFrame.
    - worker_id_col (str): The column name containing the worker IDs.
    - phone_id_col (str): The column name containing the phone IDs.
    - phone_details_cols (list): List of column names containing the phone number details to be removed.

    Returns:
    - DataFrame: The modified DataFrame with duplicate phone number details removed.
    """
    duplicates_mask = df.groupby(worker_id_col)[phone_id_col].transform(lambda x: x.duplicated())
    # Dictionary mapping column names to replacement values
    df.loc[duplicates_mask, phone_details_cols] =''

    return df

def set_primary_flag(df, worker_id_column, primary_column):
    """
    Sets the 'Primary' flag to True for the first row of each worker group if no primary phone is present.

    Args:
    - df (DataFrame): The input DataFrame.

    Returns:
    - DataFrame: The modified DataFrame with 'Primary' flag set.
    """
    # uncoment if the primary = True/ False
    #has_primary = df.groupby(worker_id_column)[primary_column].transform('any')

    # Group by worker ID and check if any row has 'Primary' set to '1'
    has_primary = df.groupby(worker_id_column)[primary_column].transform(lambda x: (x == '1').any())
    
    # Set 'Primary' flag to '1' for the first row of each group if no primary phone is present
    df.loc[~has_primary & df.groupby(worker_id_column).cumcount().eq(0), primary_column] = '1'
    
    return df

# Parse 'Full Name' into 'Legal First Name', 'Legal Middle Name', 'Legal Last Name' if 'Legal First Name' is blank
def parse_name(row):
    print(row['Legal First Name'])
    if pd.isnull(row['Legal First Name']) or row['Legal First Name'] == '':
        names = row['Full Name'].split(' ')
        row['Legal First Name'] = names[0] if len(names) >= 1 else ''
        row['Legal Last Name'] = names[-1] if len(names) >= 2 else ''
        row['Legal Middle Name'] = ' '.join(names[1:-1]) if len(names) > 2 else ''

    return row
    # Function to set primary contact as 0 for second row onwards
def set_secondary_contacts(group):
    num_primary = (group['Primary Contact'] == '1').sum()
    if num_primary >= 2:
        group['Primary Contact'].iloc[1:] = '0'

    return group

# Define a function to handle the address line wiping
def emergency_contact_wipe_address(group):
    if len(group) > 1:
        group.loc[group.index[0], 'Home Address Line 2'] = np.nan
        group.loc[group.index[1:], ['Home Address Line 1', 'City', 'Region/State', 'Postal Code']] = np.nan
        group.loc[group.index[2:], ['Home Address Line 1', 'Home Address Line 2','City', 'Region/State', 'Postal Code']] = np.nan

    return group

def personal_contact_wipe_address(group):

    # if len(group) > 1:
    #     group.loc[group.index[0], ['Home Address Line 2', 'Home Address Line 3']] = np.nan
    #     group.loc[group.index[1], ['Home Address Line 1', 'Home Address Line 3', 'City', 'Region/State', 'Postal Code']] = np.nan
    #     if 2 in group.index.values:
    #         group.loc[group.index[2], ['Home Address Line 1', 'Home Address Line 2', 'City', 'Region/State', 'Postal Code']] = np.nan
    #     if 3 in group.index.values:
    #         group.loc[group.index[3:], ['Home Address Line 1', 'Home Address Line 2', 'Home Address Line 3', 'City', 'Region/State',
    #                                     'Postal Code']] = np.nan
    if len(group) > 1:
        group.iloc[0, group.columns.get_loc('Home Address Line 2')] = np.nan
        group.iloc[0, group.columns.get_loc('Home Address Line 3')] = np.nan

        group.iloc[1, group.columns.get_loc('Home Address Line 1')] = np.nan
        group.iloc[1, group.columns.get_loc('Home Address Line 3')] = np.nan
        group.iloc[1, group.columns.get_loc('City')] = np.nan
        group.iloc[1, group.columns.get_loc('Region/State')] = np.nan
        group.iloc[1, group.columns.get_loc('Postal Code')] = np.nan

        if len(group) > 2:
            group.iloc[2, group.columns.get_loc('Home Address Line 1')] = np.nan
            group.iloc[2, group.columns.get_loc('Home Address Line 2')] = np.nan
            group.iloc[2, group.columns.get_loc('City')] = np.nan
            group.iloc[2, group.columns.get_loc('Region/State')] = np.nan
            group.iloc[2, group.columns.get_loc('Postal Code')] = np.nan

        if len(group) > 3:
            group.iloc[3:, group.columns.get_loc('Home Address Line 1')] = np.nan
            group.iloc[3:, group.columns.get_loc('Home Address Line 2')] = np.nan
            group.iloc[3:, group.columns.get_loc('Home Address Line 3')] = np.nan
            group.iloc[3:, group.columns.get_loc('City')] = np.nan
            group.iloc[3:, group.columns.get_loc('Region/State')] = np.nan
            group.iloc[3:, group.columns.get_loc('Postal Code')] = np.nan
            
    # Sort the group by index to preserve the original order
    group = group.sort_index()
    return group

def personal_contact_wipe_local_address(group):
    print("Function is being called.")
    group_length = len(group)
    print("Processing group with length:", group_length)
    print("Group indices:", group.index)
    if group_length > 0:
        group.loc[group.index[0], ['Home Address Line 2 - Local', 'Home Address Line 3 - Local']] = np.nan
    if group_length > 1:    
        group.loc[group.index[1], ['Home Address Line 1 - Local', 'Home Address Line 3 - Local', 'City - Local', 'Region/State', 'Postal Code']] = np.nan
    if group_length > 2:
        group.loc[group.index[2], ['Home Address Line 1 - Local', 'Home Address Line 2 - Local', 'City - Local','Region/State', 'Postal Code']] = np.nan
    if group_length > 3:
        group.loc[group.index[3:], ['Home Address Line 1 - Local', 'Home Address Line 2 - Local', 'Home Address Line 3 - Local', 'City - Local', 'Region/State','Postal Code']] = np.nan

    return group

def mapping_data(input_data, data_transform_mapping, mapping_file, all_unique_data_list, unavailable_reference_id_list,
                 load):
    # global all_unique_data_list  # Add global keyword
    mapping_data = pd.read_excel(mapping_file)

    # Map the values based on the mapping file and handle KeyError
    unique_values = pd.DataFrame(
        columns=['Unique Value', 'Mapped Value', 'Mapped Reference ID Type', 'Load'])

    un_mapped_values = pd.DataFrame(
        columns=['Unique Value', 'UnMapped Reference ID Type', 'Load'])

    for column, filter_value in data_transform_mapping.items():
        # input_data[column] = input_data[column].astype('str')
        # convert the column to str if it not already
        mapped_data = \
            mapping_data[mapping_data['Reference ID Type'] == filter_value].set_index('Business Object Instance')[
                'Reference ID Value']

        un_available_id_type = pd.DataFrame()
        if not len(mapped_data):
            un_mapped_values = pd.concat([
                un_mapped_values,
                pd.DataFrame({
                    'Unique Value': [column],
                    'UnMapped Reference ID Type': [filter_value],
                    'Load': [load]
                })
            ], ignore_index=True)

        input_data = pd.merge(input_data, mapped_data, left_on=column, right_on='Business Object Instance', how='left',
                              suffixes=('', '_mapped'))

        # create new dataframe with value and mapped value and collate together to form giant document
        unique_column_values = input_data[[column, 'Reference ID Value']].drop_duplicates()
        unmapped_id = 0
        # if not len(mapped_data):
        #     unmapped_id = filter_value
        unique_values = pd.concat([
            unique_values,
            pd.DataFrame({
                'Unique Value': unique_column_values[column],
                'Mapped Value': unique_column_values['Reference ID Value'],
                'Mapped Reference ID Type': filter_value,
                'Load': load
            })
        ], ignore_index=True)

        # replace the column value with mapped value in input_data dataframe

        input_data[column] = input_data['Reference ID Value']
        input_data.drop('Reference ID Value', axis=1, inplace=True)
        # input_data[column] = input_data[column].map(mapped_data)

    # # Append the unique values to the global list
    all_unique_data_list = pd.concat([all_unique_data_list, unique_values], ignore_index=True)
    unavailable_reference_id_list = pd.concat([unavailable_reference_id_list, un_mapped_values],
                                              ignore_index=True)
    # input_data.drop_duplicates(keep=False, inplace=True)
    return input_data, all_unique_data_list, unavailable_reference_id_list


def convert_data(input_data, sheet, load, mapping_file, mapping_data_dict,
                 all_unique_data_list, unavailable_reference_id_list, input_file, eib_file_name):
    start_row = 6
    columns_mapping = {}

    # CountryISO2Code = "AU"
    # CountryISO3Code = "AUS"
    # date_today = datetime.now().date().strftime('%Y-%m-%d')
    currency_code = "AUD"

    mapping_file = mapping_file
    defalut_compensation_count = 0
    if load == "Change Personal Information":
        
        input_data = input_data.dropna(subset=['Gender', 'Date of Birth', 'Marital Status', 'Marital Status Date', 'Race/Ethnicity', 'Citizenship Status', 'Citizenship Status Country', 'Pronoun',	'Gender Identity', 'Disability Status', 'Veteran Status Name'], how='all')
        input_data.reset_index(drop=True, inplace=True)
       
        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()

        # format columns to date
        convert_and_format_date(input_data, 'Date of Birth')
        convert_and_format_date(input_data, 'Marital Status Date')

        # Blank out 'Marital Status Date' column where 'Marital Status' is NaN or blank
        input_data.loc[(input_data['Marital Status'] == 'Engaged (Australia)'), 'Marital Status Date'] = ""

        columns_mapping = {
            'spreadsheet_key': 2,
            'Legacy Worker ID': 3,
            'Date of Birth': 6,
            'Gender': 9,
            'Disability Status': 14,
            'Marital Status': 36,
            'Citizenship Status': 38,
            'Citizenship Status Country': 39,
            'Race/Ethnicity': 41,
            'Marital Status Date': 72,
            'Pronoun': 82,
            'Gender Identity': 81
        }

        data_transform_mapping = {
            'Gender': 'Gender_Code',
            'Marital Status': 'Marital_Status_ID',
            'Citizenship Status': 'Citizenship_Status_Code',
            'Race/Ethnicity': 'Ethnicity_ID',
            'Disability Status': 'Disability_ID',
            'Legacy Worker ID': "Worker_WID",
            'Gender Identity': 'Gender_Identity_ID',
            'Pronoun': 'Pronoun_ID',
            'Citizenship Status Country': 'ISO_3166-1_Alpha-3_Code'
        }
        # 'Race/Ethnicity':'Ethnicity_ID','Disability Status':'Disability_ID',

        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)
        print(">>>>", len(input_data))

    elif load == "Change Legal Name":
        columns_mapping = {
            'key': 2,
            'Legacy Worker ID': 3,
            'Effective Date': 5,
            'Country ISO2 Code': 8,
            'Legal First Name': 12,
            'Legal Middle Name': 13,
            'Legal Last Name': 14
        }
        data_transform_mapping = {
            'Legacy Worker ID': "Worker_ID"
        }
        # 'Race/Ethnicity':'Ethnicity_ID','Disability Status':'Disability_ID',

        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)

    elif load == "Change Preferred Name":
        columns_mapping = {
            'key': 2,
            'Legacy Worker ID': 3,
            'Country ISO Code': 8,
            'Preferred First Name': 12,
            'Preferred Middle Name': 13,
            'Preferred Last Name': 14
        }
        data_transform_mapping = {
            'Legacy Worker ID': "Worker_ID"
        }
        # 'Race/Ethnicity':'Ethnicity_ID','Disability Status':'Disability_ID',

        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)


    elif load == "Change-Job":

        # format columns to date
        convert_and_format_date(input_data, 'Hire Date')
        convert_and_format_date(input_data, 'End Employment Date (Only for Fixed Term Employees)')

        columns_mapping = {
            'key': 2,
            'Legacy Worker ID': 3,
            'Effective Date': 5,
            'Change Job Reason': 6,
            'Employee Type': 13,
            'Job Profile': 15,
            'Business Title': 17,
            'Work Location': 18,
            'Scheduled Weekly Hours': 24,
            'Default Weekly Hours': 23,
            'Pay Rate Type (Salaried/Hourly)': 32,
            'End Employment Date (Only for Fixed Term Employees)': 43,
            'Time Type': 20
        }
        # additional mapping that needs to be done
        data_transform_mapping = {
            'Employee Type': 'Employee_Type_ID',
            'Job Profile': 'Job_Profile_ID',
            'Work Location': 'Location_ID',
            'Pay Rate Type (Salaried/Hourly)': 'Pay_Rate_Type_ID',
            'Time Type': 'Time_Type_ID',
            'Legacy Worker ID': "Worker_ID"
        }

        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)

    elif load == "Change Job Assign Pay Group":

        columns_mapping = {
            'key': 2,
            'Pay Group Name': 3,
        }

        data_transform_mapping = {'Pay Group Name': "Organization_Reference_ID",'Legacy Worker ID': "Worker_ID"}
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)

    elif load == "Legacy ID":

        columns_mapping = {
            'key': 2,
            'Worker ID': 3,
            'Legacy Worker ID': 9
        }

    elif load in ["Service Dates", "Contingent Service Dates"]:
        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()
        from get_mapping_for_loads import get_edit_service_dates_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_edit_service_dates_data(input_data, mapping_file,
                                        all_unique_data_list,
                                        unavailable_reference_id_list, load)
       
    elif load == "National ID":

        # format columns to date
        #remove_non_numeric_char(input_data, "National ID")

        input_data['spreadsheet_key'] = (
                input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()

        input_data = generate_row_id(input_data, 'spreadsheet_key', "National ID", "Row ID")

        columns_mapping = {
            'spreadsheet_key': 2,
            'Legacy Worker ID': 3,
            'National ID Country': 11,
            'National ID Type Name': 10,
            'National ID': 9,
            'Row ID': 6,
            'Issuing Agency': 16,
            'Issued Date': 12,
            'Expiration Date': 13

        }

        # additional mapping that needs to be done
        data_transform_mapping = {#'National ID Country': 'ISO_3166-1_Alpha-3_Code',
                                  'National ID Type Name': 'National_ID_Type_Code','Legacy Worker ID': "Worker_WID"}
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)

    elif load == "Visas":
        print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>', input_data['Expiration Date'])
        # format columns to date
        convert_and_format_date(input_data, 'Issued Date')
        convert_and_format_date(input_data, 'Expiration Date')
        print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>', input_data['Expiration Date'])

        columns_mapping = {
            'key': 2,
            'Legacy Worker ID': 3,
            'Visa ID': 20,
            'Visa Type ID': 21,
            'Country ISO Code': 22,
            'Issued Date': 23,
            'Expiration Date': 24,
        }
        # additional mapping that needs to be done
        data_transform_mapping = {'Visa Type ID': 'Visa_ID_Type_ID','Legacy Worker ID': "Worker_WID"}
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)

    elif load == "Passports":
        # format columns to date
        convert_and_format_date(input_data, 'Issued Date')
        convert_and_format_date(input_data, 'Expiration Date')

        columns_mapping = {
            'key': 2,
            'Legacy Worker ID': 3,
            'Passport ID': 9,
            'Passport Type': 10,
            'Country ISO3 Code': 11,
            'Issued Date': 12,
            'Expiration Date': 13
        }
        # additional mapping that needs to be done
        data_transform_mapping = {'Passport Type': 'Passport_ID_Type_ID','Legacy Worker ID': "Worker_WID"}
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)

    elif load == "Payment Elections":

        # replace 0 with blanks for amount and pct field
        input_data['Distribution Amount*'] = input_data['Distribution Amount*'].replace(0, "")
        input_data['Distribution Percentage*'] = input_data['Distribution Percentage*'].replace(0, "")

        # fill the bank identification code if it is missing 0

        input_data['Bank Identification Code'] = input_data['Bank Identification Code'].astype(str).str.replace(r'\.0$',
                                                                                                                '',
                                                                                                                regex=True)
        input_data['Bank Identification Code'] = input_data['Bank Identification Code'].astype(str).str.zfill(6)
        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()

        columns_mapping = {
            'spreadsheet_key': 2,
            'Legacy Worker ID': 7,
            'Country': [8, 15, 18],
            'Bank Sort Code': [12, 13],
            'Currency*': [9, 16, 19],
            'Account Nick Name':20,
            'Name On Account': 21,
            'Bank Account No.': 22,
            'Roll Number': 23,
            'Account Type': 25,
            'Bank Name': 26,
            'IBAN Number': 27,
            'Bank Identification Code': 28,
            'Branch Name': 30,
            'Distribution Amount*': 33,
            'Distribution Percentage*': 34,
            'Distribution Balance*': 35,
        }

        data_transform_mapping = {'Account Type': 'Bank_Account_Type', 'Distribution Balance*': "Numeric_Boolean_ID",'Legacy Worker ID': "Worker_WID"}
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)



    # elif load == "Work Contact Change":

    #     remove_non_numeric_char(input_data, "Work Phone Number")
    #     columns_mapping = {
    #         'Work Phone Number': 56,
    #         'Work Email': 72
    #     }

    elif load == "Change Job Change Org":

        # for column in input_data.columns:
        #     if input_data[column].dtype == 'float64':
        #         is_valid = np.isfinite(input_data[column])  # Check for finite values (not NaN or inf)
        #         input_data.loc[is_valid, column] = input_data.loc[is_valid, column].astype(str).str.rstrip('.0')

        input_data['Current Cost Center Code'] = input_data['Current Cost Center Code'].astype(str).str.replace(r'\.0$',
                                                                                                                '',
                                                                                                                regex=True)
        input_data['Cost Center Code'] = input_data['Cost Center Code'].astype(str).str.replace(r'\.0$', '', regex=True)

        # Select the rows where current value for attribute is equal to Future value
        mask_business_area = input_data['Business Area'] == input_data['Current Business Area']
        mask_division = input_data['Division'] == input_data['Current Division']
        mask_cost_center = input_data['Cost Center Code'] == input_data['Current Cost Center Code']

        # Update the columns with np.nan values for the selected rows
        input_data.loc[mask_business_area, ['Business Area', 'Current Business Area']] = np.nan
        input_data.loc[mask_division, ['Division', 'Current Division']] = np.nan
        input_data.loc[mask_cost_center, ['Cost Center Code', 'Current Cost Center Code']] = np.nan

        # columns that i need to convert to a row
        columns_to_melt = ['Business Area', 'Cost Center Code', 'Division', 'Current Business Area',
                           'Current Cost Center Code', 'Current Division']
        # define the key
        id_vars = ['Legacy Worker ID', 'Company Name']

        var_name = "Org_Type"
        value_name = "Org"

        input_data = convert_column_to_row(input_data, id_vars, columns_to_melt, var_name, value_name)
        # input_data.to_csv('Beforemerge.csv')

        if load == "Change Job Change Org":
            # read the change Job file as we will need to rowIds to match between change job and this load
            change_job_file = base_path + 'EIB Template\\Change_Job_v40.1.xlsx'
            change_job_sheet = "Change Job"

            additional_data = pd.read_excel(change_job_file, sheet_name=change_job_sheet, skiprows=[0, 1, 2, 3])

            additional_data = additional_data[["Spreadsheet Key*", "Worker*"]]

            input_data = pd.merge(input_data, additional_data, how='left', left_on="Legacy Worker ID",
                                  right_on="Worker*").sort_values('Spreadsheet Key*')

        # Add 'Row ID' and 'Delete' columns
        input_data['Row ID'] = input_data[var_name].apply(lambda x: 1 if 'Current' in x else 2)
        input_data['Delete'] = input_data[var_name].apply(lambda x: 'Y' if 'Current' in x else 'N')

        # input_data.sort_values(by=['Spreadsheet Key*',"Row ID"], inplace = True)

        columns_mapping = {
            value_name: 8,
            'Spreadsheet Key*': 2,
            'Company Name': 3,
            'Row ID': 6,
            'Delete': 7,
        }

        # additional mapping that needs to be done
        data_transform_mapping = {value_name: 'Custom_Organization_Reference_ID',
                                  'Company Name': 'Company_Reference_ID','Legacy Worker ID': "Worker_ID"}  # value_name:'Custom_Organization_Reference_ID'
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)

    elif load == "Absence Input":
        # format columns to date

        convert_and_format_date(input_data, 'Next period Start Date')
        # input_data['vacation_plan_start'] = "2023-01-01"
        # input_data['vacation_plan_end'] = "2023-12-31"
        # input_data['snapshot_date'] = "2023-09-30"

        columns_mapping = {
            'key': 2,
            'Legacy Worker ID': 6,
            'Reference ID of Time off Plan': 8,
            'Balance as of Period Date': 10,
            'Next period Start Date': 9,
            # 'vacation_plan_start': 9,
            # 'vacation_plan_end': 10,
            # 'snapshot_date': 11,
            # 'Vacation Entitlement': 12,
        }

        # additional mapping that needs to be done
        data_transform_mapping = {'Reference ID of Time off Plan': 'Absence_Plan_ID', 'Legacy Worker ID': "Worker_ID"}
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)


    elif load == "Emergency Contact":
        # format columns to date
        remove_non_numeric_char(input_data, "Emergency Contact Primary Home Phone Number")
        remove_non_numeric_char(input_data, "Emergency Contact Primary Work Phone Number")
        remove_non_numeric_char(input_data, "Emergency Contact Additional First Home Phone")
        remove_non_numeric_char(input_data, "Emergency Contact Additional Second Home Phone")
        remove_non_numeric_char(input_data, "Emergency Contact Additional First Work Phone")
        remove_non_numeric_char(input_data, "Emergency Contact Additional Second Work Phone")

        #input_data = input_data[input_data['Legacy Worker ID'].isin(["705128", "261204"])]
        input_data = input_data.sort_values(by=['Legacy Worker ID', 'Primary Contact','Full Name'], ascending=[True, False, True]).reset_index(drop=True)
        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()
        # Changing the 'Full Name' column to string type
        input_data['Full Name'] = df['Full Name'].astype(str)
        #parsing full name for emergency contact that have no names
        input_data = input_data.apply(parse_name, axis=1)
        # Apply the function to each group of workers
        input_data = input_data.groupby('Legacy Worker ID', as_index=False).apply(set_secondary_contacts).reset_index(drop=True)

        
        #DATA ISSUE: WIPE OUT PHONE NUM IF THEY ARE SAME For PRIMARY Vs ADDITIONAl
        input_data['Emergency Contact Additional First Home Phone'] = input_data.apply(lambda x: np.nan if x['Emergency Contact Primary Home Phone Number'] == x['Emergency Contact Additional First Home Phone'] else x['Emergency Contact Additional First Home Phone'], axis=1)
        input_data['Emergency Contact Additional First Work Phone'] = input_data.apply(lambda x: np.nan if x['Emergency Contact Primary Work Phone Number'] == x['Emergency Contact Additional First Work Phone'] else x['Emergency Contact Additional First Work Phone'], axis=1)
        input_data = input_data.replace('',np.nan)
        # columns that i need to convert to a row
        columns_to_melt = ['Emergency Contact Primary Home Phone Number', 'Emergency Contact Primary Work Phone Number',
                           'Emergency Contact Additional First Home Phone', 'Emergency Contact Additional Second Home Phone',
                           'Emergency Contact Additional First Work Phone',
                           'Emergency Contact Additional Second Work Phone']
        
        input_data['Number of Phone'] = input_data.groupby(['Legacy Worker ID','Emergency Contact ID'])[columns_to_melt].transform(lambda x: x.count()).sum(axis=1)
        #print(input_data[['Home Address Line 1', 'Home Address Line 2','Number of Phone']])
        # define the key
        id_vars = ['spreadsheet_key', 'Legacy Worker ID', 'Worker ID','Emergency Contact ID','Primary Contact', 'Priority', 'Country',
                   'Full Name','Legal First Name', 'Legal Middle Name','Legal Last Name', 
                   'Relationship', 'Prefered Language', 'Home Address Line 1', 'Home Address Line 2',
                   'City', 'Region/State', 'Postal Code', 'Home Email','Number of Phone']

        var_name = "Phone Type"
        value_name = "Phone Numbers"

        input_data = convert_column_to_row(input_data, id_vars, columns_to_melt, var_name, value_name)
        
        #remove unwanted data
        input_data = input_data.groupby(['Legacy Worker ID', 'Primary Contact', 'Emergency Contact ID'], sort=False).apply(
        emergency_contact_wipe_address).reset_index(drop=True)
        
        input_data = input_data.sort_values(by=['Legacy Worker ID', 'Primary Contact','Full Name'], ascending=[True, False, True]).reset_index(drop=True)
        # print(input_data[['Home Address Line 1', 'Home Address Line 2']])
        input_data = generate_row_id(input_data, ['spreadsheet_key','Emergency Contact ID'], "Phone Type", "Phone Row ID")
        input_data = generate_row_id(input_data, 'spreadsheet_key', 'Emergency Contact ID', "Contact Row ID")
        
        input_data['Default_value_1'] = '1'
        input_data['Default_value_N'] = 'N'
        input_data['Primary Contact'] = np.where(input_data['Primary Contact'] == '1', 'Y', 'N')
        input_data['Phone Device Type'] = input_data.apply(lambda row: '' if (pd.isna(row['Phone Numbers']) or  row['Phone Numbers']=='') else "PHONE_DEVICE_TYPE_Home" if "Home" in row['Phone Type'] else "PHONE_DEVICE_TYPE_Business", axis=1)
        input_data['Phone Type Primary'] = input_data.apply(lambda row: '' if (pd.isna(row['Phone Numbers']) or  row['Phone Numbers']=='')  else 'Y' if row['Number of Phone'] == 1 else "Y" if "Primary" in row['Phone Type'] else "N", axis=1)
        input_data['Phone Communication Usage Type'] = input_data.apply(lambda row: '' if (pd.isna(row['Phone Numbers']) or row['Phone Numbers']=='')  else "HOME" if "Home" in row['Phone Type'] else "WORK", axis =1)
        
        # print(input_data[['Legacy Worker ID','Emergency Contact ID',"Contact Row ID",'Primary Contact',"Phone Numbers"]])

        # setting priority = 2 when primary contact = 0 and priority =1
        input_data.loc[(input_data['Primary Contact'] == "N") & (input_data['Priority'] == "1"), 'Priority'] = "2"

        # setting primary Y when phone row id =1 and primary - N
        input_data.loc[(input_data["Phone Row ID"] == 1) & (input_data['Phone Type Primary'] == "N"), 'Phone Type Primary'] = "Y"

        columns_mapping = {
            'spreadsheet_key': 2,
            'Legacy Worker ID': 3,
            'Primary Contact':9,
            'Priority': 10,
            'Contact Row ID': [5,106],
            'Legal First Name': 20,
            'Legal Middle Name': 21,
            'Legal Last Name': 22,
            'Relationship': 11,
            'Prefered Language': 12,
            'Home Address Line 1': 119,
            'City': 120,
            'Region/State': 127,
            'Country': [16,156],
            'Postal Code': 133,
            "Phone Row ID":146,
            'Default_value_1': [13,161,163],
            'Phone Device Type': 160,
            'Phone Type Primary': 164,
            'Phone Communication Usage Type': 165,
            #'Phone Type': 160,
            'Default_value_N': 162,
            'Home Email': 174,
            'Phone Numbers': 158,
            #'Home Phone Number Country Code': 157
        }
        print('>>>>>>>>>>>>>>>>>', len(input_data))
        # additional mapping that needs to be done
        data_transform_mapping = {'Region/State': 'Country_Region_ID',
                                  'Relationship': 'Related_Person_Relationship_ID', #'Prefered Language': 'Language_ID',
                                  'Legacy Worker ID': "Worker_WID"}
    
        input_data, all_unique_data_list, unavailable_reference_id_list = mapping_data(input_data,
                                                                                       data_transform_mapping,
                                                                                       mapping_file,
                                                                                       all_unique_data_list,
                                                                                       unavailable_reference_id_list,
                                                                                       load)
        input_data.drop_duplicates(inplace=True)
        input_data.reset_index(drop = True, inplace=True)
        #input_data = input_data[input_data["Legacy Worker ID"].isin(["ab45f523e3b01001c63b59f2edbc0000","fb9704e1204910049347ba7e3f860000"])]


    elif load == "Supervisory Org":
        from get_mapping_for_loads import get_mapped_supervisory_org_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_supervisory_org_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        #Sorting the data by Organization Level from Top column first row come by the levels
        input_data = input_data.sort_values(by=['Organization Level from Top'], ignore_index=True)


    elif load == "Job Requisitions":     
        from get_mapping_for_loads import get_mapped_job_requisitions_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_requisitions_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        

    elif load == "Create Position":
        
        # we need to re visit the below line while processing on P1 Build
        input_data['spreadsheet_key'] = (input_data['Position ID'] != input_data['Position ID'].shift()).cumsum()
        temp_input_data = input_data[["spreadsheet_key","Organization_Reference_ID", "Position_Request_Reason_ID", "Position ID", "Job Posting Title",
                                         "Availability Date", "Earliest Hire Date", "Job_Profile_ID", "Location_ID", "Worker_Type_ID",
                                         "Position_Time_Type_ID", "Worker Sub-Type", "Scheduled Weekly Hours", "Default Weekly Hours",
                                         "Company Name", "Division", "Business Area", "Department", "Industrial Code", "Is Overlapped Position"]]
        temp_input_data.drop_duplicates(inplace=True)
        
        #hardcoding for failing job profile missing title
        temp_input_data["Job Posting Title"] = temp_input_data['Job Posting Title'].fillna("Tech Lead-Site Production Platform (SPP)")

        if sheet.title == 'Create Position':
            from get_mapping_for_loads import get_mapped_create_position_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_create_position_data(
                temp_input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)

        elif sheet.title == 'Request Default Compensation':
            
            #replace the "0" for allowance columns only
            input_data[['Allowance Amount', 'Allowance Percentage']] = input_data[['Allowance Amount', 'Allowance Percentage']].replace("0", "")
            #overwriting the 0 salary with 0.1
            # input_data.loc[(input_data['Salary Amount'] == '0') & (input_data['Salary Plan'] != ''), 'Salary Amount'] = '0.1'
            
            #input_data.loc[(input_data['Allowance Amount'] == '0') & (input_data['Allowance Plan'] == 'Casual Loading %'), 'Allowance Percentage'] = '0.001'

            from get_mapping_for_loads import get_request_default_compensation
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_request_default_compensation(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list, eib_file_name)

        elif sheet.title == 'Edit Assign Organization':

            from get_mapping_for_loads import get_assign_org_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_assign_org_data(
                temp_input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list, eib_file_name)

    elif load in ["Prehire", "Future Prehire"]:
        #input_data.dropna(subset=['Email Address'], inplace=True)
        input_data['Email Address'] = input_data['Email Address'].fillna("dummyemail@forload.com")
        # we need to re visit the below line while processing on P1 Build
       
        from get_mapping_for_loads import get_mapped_prehire_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_prehire_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        input_data['Row_Count'] = input_data.groupby('Applicant ID').cumcount() + 1


    elif load in ['Hire Employee', 'Overlapping Hire Employee', 'Future Hires']:
        # Add a counter column
        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()

        if sheet.title == 'Hire Employee':
            from get_mapping_for_loads import get_mapped_hire_employee_data

            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_hire_employee_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)
            
        elif sheet.title == 'Assign Pay Group':
            from get_mapping_for_loads import get_mapped_hire_employee_assign_pay_group_data

            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_hire_employee_assign_pay_group_data(
                input_data, mapping_file, all_unique_data_list, unavailable_reference_id_list, eib_file_name)
        
        elif sheet.title == 'Edit Assign Organization':

            from get_mapping_for_loads import get_hire_assign_org_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_hire_assign_org_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list, eib_file_name)
            
        elif sheet.title == 'Edit Service Dates':

            from get_mapping_for_loads import get_edit_service_dates_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_edit_service_dates_data(
                                                input_data, mapping_file, all_unique_data_list, unavailable_reference_id_list, load)

    elif load in ['Hire CWR', 'Future Hire CWR']:
        
        if sheet.title == 'Contract Contingent Worker':
            from get_mapping_for_loads import get_mapped_hire_cwr_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_hire_cwr_data(
                input_data, mapping_file, all_unique_data_list, unavailable_reference_id_list)
        
        elif sheet.title == 'Edit Assign Organization':

            from get_mapping_for_loads import get_hire_cwr_assign_org_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_hire_cwr_assign_org_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list, eib_file_name)
            
        elif sheet.title == 'Edit Service Dates':

            from get_mapping_for_loads import get_edit_service_dates_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_edit_service_dates_data(
                                            input_data, mapping_file, all_unique_data_list, unavailable_reference_id_list, load)

    elif load == 'Change Job':
        # Add a counter column
        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()

        if sheet.title == 'Change Job':
            from get_mapping_for_loads import get_mapped_change_job_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_change_job_data(
                input_data, mapping_file,
                all_unique_data_list, unavailable_reference_id_list)
            
        elif load == 'Change Organization':
            from get_mapping_for_loads import get_mapped_change_job_change_organization_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_change_job_change_organization_data(
                input_data, mapping_file,
                all_unique_data_list, unavailable_reference_id_list)

        elif load == 'Assign Pay Group':
            from get_mapping_for_loads import get_mapped_change_job_assign_pay_group_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_change_job_assign_pay_group_data(
                input_data, mapping_file,
                all_unique_data_list, unavailable_reference_id_list) 

    elif load == 'Terminate Employee':
        from get_mapping_for_loads import get_mapped_terminate_emp_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_terminate_emp_data(
            input_data, mapping_file, all_unique_data_list, unavailable_reference_id_list)

    elif load == 'End Contingent Worker Contracts':
        from get_mapping_for_loads import get_mapped_end_contingent_worker_contracts_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_end_contingent_worker_contracts_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'One Time Payments':
        from get_mapping_for_loads import get_mapped_one_time_payments_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_one_time_payments_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Compensation':
        from get_mapping_for_loads import get_mapped_compensation_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_compensation_data(
            input_data, mapping_file, all_unique_data_list, unavailable_reference_id_list)

        from get_mapping_for_loads import make_compensation_eib_data_sheet
        sheet, all_unique_data_list, unavailable_reference_id_list = make_compensation_eib_data_sheet(input_data, sheet,
                                                                                                      all_unique_data_list,
                                                                                                      unavailable_reference_id_list)
        return sheet, all_unique_data_list, unavailable_reference_id_list

    elif load == 'Put Candidate':

        remove_non_numeric_char(input_data, "Phone Number")

        from get_mapping_for_loads import get_mapped_put_candidate_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_put_candidate_data(
            input_data, mapping_file,
            all_unique_data_list, unavailable_reference_id_list)
    
    elif load == 'Job Requisition Roles':
        input_data['spreadsheet_key'] = (input_data['Job Requisition ID'] != input_data['Job Requisition ID'].shift()).cumsum()
        from get_mapping_for_loads import get_mapped_role_based_requisition_roles_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_role_based_requisition_roles_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)

    elif load == 'Edit Job Requisitions':
        input_data['spreadsheet_key'] = (input_data['Job Requisition ID'] != input_data['Job Requisition ID'].shift()).cumsum()
        if sheet.title == 'Edit Job Requisition':
            from get_mapping_for_loads import get_mapped_job_requisition_roles_edit_job_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_requisition_roles_edit_job_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)
            
        elif sheet.title == 'Assign Organization Roles':
            from get_mapping_for_loads import get_mapped_job_requisition_roles_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_requisition_roles_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)

    elif load == 'Performance Reviews':
        input_data['spreadsheet_key'] = (input_data['Legacy Worker ID'] != input_data['Legacy Worker ID'].shift()).cumsum()
        if sheet.title == 'Start Performance Review':
            from get_mapping_for_loads import get_mapped_start_performance_reviews_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_start_performance_reviews_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)
            
        elif sheet.title == 'Complete Manager Evaluation':
            from get_mapping_for_loads import get_mapped_performance_manager_evaluation_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_performance_manager_evaluation_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)

    elif load == 'Probation Info':
        from get_mapping_for_loads import get_mapped_probation_info_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_probation_info_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Manage Goals':
        from get_mapping_for_loads import get_mapped_manage_goals_data
    
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_manage_goals_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Supplier':
        from get_mapping_for_loads import get_mapped_supplier_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_supplier_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Assign Work Schedule':
        from get_mapping_for_loads import get_mapped_assign_work_schedule_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_assign_work_schedule_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Leave Of Absence Events':
        
        from get_mapping_for_loads import get_mapped_leave_of_absence_event_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_leave_of_absence_event_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Add Workday Account':
        input_data = input_data.replace({True: 'Y', False: 'N'})
        from get_mapping_for_loads import get_mapped_add_workday_event_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_add_workday_event_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Role Based Assignments':
        #need to add suffix if Supervisory Org
        input_data.loc[input_data['Organization Type'] == 'Supervisory', 'Role Assigner*'] = 'NCA_' + input_data['Role Assigner*']
        
        #For Organization partner overwrite assigning org to NCA division
        input_data.loc[input_data['Assignable Role*'] == 'Organization_Partner', 'Role Assigner*'] = 'DIVISION_NCA'
        input_data.loc[input_data['Assignable Role*'] == 'Recruiting_Coordinator', 'Role Assigner*'] = 'DIVISION_NCA'
        input_data.loc[input_data['Assignable Role*'] == 'HR_Director', 'Role Assigner*'] = 'DIVISION_NCA'
        input_data.loc[input_data['Assignable Role*'] == 'NCA_Payroll_Partner', 'Role Assigner*'] = 'DIVISION_NCA'
        input_data.loc[input_data['Assignable Role*'] == 'NCA_Payroll_Interface_Partner', 'Role Assigner*'] = 'DIVISION_NCA'
        
        #create extra record for recruting coordinator as this will need to be applied additional role
        rec_coordinator = input_data[input_data['Assignable Role*'] == 'Recruiting_Coordinator']
        rec_coordinator.loc[:,'Assignable Role*'] = 'ASSIGNABLE_ROLE-3-135'
        rec_coordinator.loc[:,'Role Assigner*'] = 'LOCATION_APAC'#assign to APAC Location Hierarchy

        #create extra record for General Manager of Finance role as this will need to be applied additional role in NC Tenant
        finance_partner = input_data[input_data['Assignable Role*'] == 'General_Manager_of_Finance']
        finance_partner.loc[:,'Assignable Role*'] = 'ASSIGNABLE_ROLE-3-135' #head of location
        finance_partner.loc[:,'Role Assigner*'] = 'LOCATION_APAC'#assign to APAC Location Hierarchy
        input_data =  pd.concat([input_data,rec_coordinator, finance_partner], ignore_index=True)

        input_data.reset_index(drop=True,inplace=True)
        #divide the role based file to two dataframe
        input_data_sup = input_data[input_data['Assignable Role*'].isin(['Manager','Workforce_Planner', 'Manager_Plus','ASSIGNABLE_ROLE-3-132'])].reset_index(drop=True) #
       
        input_data_non_sup = input_data[~input_data['Assignable Role*'].isin(['Matrix_Manager','Manager','Manager_Plus','Workforce_Planner', 'ASSIGNABLE_ROLE-3-132', 
                                                                            'ASSIGNABLE_ROLE-3-133','Chief_Financial_Officer'])].reset_index(drop=True) #Chief_Financial_Officer
        #input_data_sup = input_data[input_data['Assignable Role*'].isin(['Manager', 'Manager_Plus'])].reset_index(drop=True) #
       
        #input_data_non_sup = input_data[~input_data['Assignable Role*'].isin(['Manager','Manager_Plus'])].reset_index(drop=True)
        from get_mapping_for_loads import get_mapped_role_based_assignments_data
        input_data_sup, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_role_based_assignments_data(
            input_data_sup, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list,'Supervisory')
       
        input_data_non_sup, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_role_based_assignments_data(
            input_data_non_sup, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list,'Non-Supervisory')
       
        # After processing, join the two halves back together
        input_data = pd.concat([input_data_sup, input_data_non_sup], ignore_index=True)
        input_data.drop_duplicates(inplace=True)
        print(len(input_data))

        input_data.sort_values(by='Position ID', inplace=True)
        input_data.reset_index(drop=True,inplace=True)
        #create spreadsheet_key after joing the data
        input_data['spreadsheet_key'] = (input_data['Position ID'] != input_data['Position ID'].shift()).cumsum()
        #input_data = generate_row_id(input_data, 'spreadsheet_key', "Assignable Role*", "Role Row ID")
        input_data['Role Row ID'] = input_data.groupby('spreadsheet_key').cumcount() + 1


    elif load == 'Worker Additional Data':

        from get_mapping_for_loads import get_mapped_edit_worker_additional_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_edit_worker_additional_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'NCA Additional Data Position':

        from get_mapping_for_loads import get_mapped_nca_additional_data_position
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_nca_additional_data_position(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Edit Position Restrictions Additional Data':
        from get_mapping_for_loads import get_mapped_edit_position_restrictions_additional_data_position
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_edit_position_restrictions_additional_data_position(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'International Assignments':
        input_data = input_data.sort_values('Effective Date').groupby('Legacy Worker ID').tail(1)
        if sheet.title == 'Add Additional Job':
            from get_mapping_for_loads import get_mapped_international_add_additional_job_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_international_add_additional_job_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list)
            
        elif sheet.title == 'Edit Assign Organization':
            from get_mapping_for_loads import get_mapped_international_edit_assign_org_data
            input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_international_edit_assign_org_data(
                input_data, mapping_file,
                all_unique_data_list,
                unavailable_reference_id_list
            )
    
    elif load == 'Company':
        from get_mapping_for_loads import get_mapped_company_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_company_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
            
    elif load == 'Job Family':
        from get_mapping_for_loads import get_mapped_job_family_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_family_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == 'Job Family Group':
        
        from get_mapping_for_loads import get_mapped_job_family_group_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_family_group_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        input_data = input_data.drop_duplicates()
        input_data = input_data.dropna(subset=['Job Family Group Name'])
        input_data = input_data.sort_values(by=['Job Family Group ID'], ignore_index=True)
   
    elif load == 'Cost Center':
        from get_mapping_for_loads import get_mapped_cost_center_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_cost_center_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Location Hierarchy":
        from get_mapping_for_loads import get_mapped_location_hierarchy_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_location_hierarchy_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        
    elif load == "Custom Organizations":
        #override Superior org with Org Ref Id
        input_data['Superior Organization Name'] = input_data['Organization ID']
        #input_data = input_data[input_data['Organization Subtype'] == "Industrial Code"]
        input_data = input_data.sort_values(by=['Organization Subtype'],ascending=True, ignore_index=True)

        from get_mapping_for_loads import get_mapped_custom_organizations_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_custom_organizations_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        
        
    elif load == "Cost Center Hierarchy":
        from get_mapping_for_loads import get_mapped_cost_center_hierarchy_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_cost_center_hierarchy_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        input_data = input_data.sort_values(by=['Org Level from Top'], ignore_index=True)

    elif load == "Collective Agreement":
        from get_mapping_for_loads import get_mapped_collective_agreement_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_collective_agreement_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Job Classification":
        from get_mapping_for_loads import get_mapped_job_classification_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_classification_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        
        input_data = input_data.sort_values(by=['Job Classification Group Id'], ignore_index=True)

    elif load == "Job Requisition Additional Data":
        from get_mapping_for_loads import get_mapped_job_requisition_additional_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_requisition_additional_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        
    elif load == "Put Supervisory Assignment Restrictions":
        from get_mapping_for_loads import get_mapped_put_supervisory_assignment_restrictions_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_put_supervisory_assignment_restrictions_data(
            input_data, columns_mapping,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Location":
        from get_mapping_for_loads import get_mapped_location_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_location_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Job Category":
        from get_mapping_for_loads import get_mapped_job_category_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_category_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Job Profile":
        from get_mapping_for_loads import get_mapped_job_profile_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_profile_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Comp Grade and Grade Profile":
    
        from get_mapping_for_loads import get_mapped_comp_grade_and_grade_profile_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_comp_grade_and_grade_profile_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load in ["Employee Compensation", "Overlapping Employee Compensation", "Future Comp Change"]:
        
        from get_mapping_for_loads import get_mapped_employee_compensation_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_employee_compensation_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load in ["Update Workday Account"]:
        from get_mapping_for_loads import get_mapped_edit_workday_account_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_edit_workday_account_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Time Off Events":
        from get_mapping_for_loads import get_mapped_time_off_events_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_time_off_events_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Assign Notice Period":
        from get_mapping_for_loads import get_mapped_assign_notice_period_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_assign_notice_period_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "User Based Assignments":
        from get_mapping_for_loads import get_mapped_user_based_assignments_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_user_based_assignments_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        
    elif load == "Licenses":
        from get_mapping_for_loads import get_mapped_licenses_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_licenses_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Job History":
        from get_mapping_for_loads import get_mapped_job_history_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_history_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Education":
        from get_mapping_for_loads import get_mapped_education_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_education_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Other Ids":
        from get_mapping_for_loads import get_mapped_other_ids_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_other_ids_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
        
    elif load == "Skills":

        from get_mapping_for_loads import get_mapped_skills_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_skills_data(input_data, mapping_file,
                                           all_unique_data_list,
                                           unavailable_reference_id_list)

    elif load == "Flexible Work Arrangements":
        from get_mapping_for_loads import get_mapped_flexible_work_arrangements_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_flexible_work_arrangements_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Worker Collective Agreement":
        from get_mapping_for_loads import get_mapped_worker_collective_agreement_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_worker_collective_agreement_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Personal Contact Change":
        from get_mapping_for_loads import get_mapped_worker_personal_contact_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_worker_personal_contact_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Work Contact Change":
        from get_mapping_for_loads import get_mapped_worker_work_contact_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_worker_work_contact_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Job History Company":
        from get_mapping_for_loads import get_mapped_job_history_company_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_job_history_company_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Pronoun Public Preference":
        from get_mapping_for_loads import get_mapped_pronoun_preferences_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_pronoun_preferences_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Future Termination":
        from get_mapping_for_loads import get_mapped_future_termination_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_future_termination_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    elif load == "Future Hires - Old":
        from get_mapping_for_loads import get_mapped_future_hires_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_future_hires_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Future Job Change":
        from get_mapping_for_loads import get_mapped_future_job_change_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_future_job_change_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Future Comp Change-":
        from get_mapping_for_loads import get_mapped_future_comp_change_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_future_comp_change_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "Skills Reference Data":
        from get_mapping_for_loads import get_mapped_skills_reference_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_skills_reference_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)
    
    elif load == "End CWR Contract":
        from get_mapping_for_loads import get_mapped_end_cwr_contract_data
        input_data, columns_mapping, all_unique_data_list, unavailable_reference_id_list = get_mapped_end_cwr_contract_data(
            input_data, mapping_file,
            all_unique_data_list,
            unavailable_reference_id_list)

    row_count = 1
    position_start_row = 6
    last_worker = 0
    last_index = 0
    emergency_priority = 1
    job_req_validation = {}
    temp_dict = {}
    comp_grad_count, comp_profile_count, comp_step_count = 1, 1, 1

    if sheet.title == 'Request Default Compensation':
        last_index = defalut_compensation_count

    # input_data = input_data.head(100)
    #input_data = input_data[input_data["Legacy Worker ID"]=="5f2619b812681001715be7c1bebc0000"]
    
    
    for idx, row in input_data.iterrows():
        if row.get('Legacy Worker ID'):
            if ((load != "Personal Work Contact Change") or (
                    load == "Personal Work Contact Change" and pd.notna(row['Work Phone Number']))):# or (
                    # load == "Work Contact Change" and pd.notna(row['Work Email']))):
                
                if last_worker == row['Legacy Worker ID']:
                    last_index = last_index
                    emergency_priority += 1
                else:
                    last_index += 1
                    emergency_priority = 1

                for column, cell_value in columns_mapping.items():
                    #write for spreadsheet key
                    if column in ['key', 'Spreadsheet Key*']:
                        sheet.cell(row=start_row, column=cell_value, value=last_index)

                    elif column == 'RowID':
                        if isinstance(cell_value, list):
                            cell_value_list = cell_value
                            for i, value in enumerate(cell_value_list):
                                sheet.cell(row=start_row, column=value, value=1)
                        else:
                            sheet.cell(row=start_row, column=cell_value, value=1)

                    elif column == 'Change Job Reason':
                        sheet.cell(row=start_row, column=cell_value, value="SUB_CATEGORY_Chg - Company")

                    elif column in ('Country ISO Code', 'Country ISO3 Code', 'Country ISO2 Code'):
                        if isinstance(cell_value, list):
                            cell_value_list = cell_value
                            for i, value in enumerate(cell_value_list):
                                sheet.cell(row=start_row, column=value, value=row[column])
                        else:
                            sheet.cell(row=start_row, column=cell_value, value=row[column])
                    
                    elif column in ('Bank Sort Code'):
                        if isinstance(cell_value, list):
                            cell_value_list = cell_value
                            for i, value in enumerate(cell_value_list):
                                sheet.cell(row=start_row, column=value, value=row[column])

                    # elif column == 'Effective Date':
                    #     sheet.cell(row=start_row, column=cell_value, value=date_today)

                    elif column in ('Currency', 'Compensation Currency'):
                        if isinstance(cell_value, list):
                            cell_value_list = columns_mapping[column]  # Get the list of values for the key
                            for i, cell_value in enumerate(cell_value_list):
                                sheet.cell(row=start_row, column=cell_value, value=row[column])
                        else:
                            sheet.cell(row=start_row, column=cell_value, value=row[column])

                    # elif load == "Work Contact Change" and (
                    #         pd.notna(row['Work Phone Number'])):# or pd.notna(row['Work Email'])):
                    #     sheet.cell(row=start_row, column=4, value=date_today)
                    #     sheet.cell(row=start_row, column=2, value=last_index)
                    #     sheet.cell(row=start_row, column=3, value=row['Legacy Worker ID'])

                    elif load == 'Compensation':
                        pass

                    # elif load == "Create Position":
                    #     if sheet.title == 'Edit Assign Organization':
                    #         pass

                    else:
                        if isinstance(cell_value, list):
                            cell_value_list = cell_value
                            for i, value in enumerate(cell_value_list):
                                sheet.cell(row=start_row, column=value, value=row[column])
                        else:
                            sheet.cell(row=start_row, column=cell_value, value=row[column])

        else:
            last_index += 1
            emergency_priority = 1
            for column, cell_value in columns_mapping.items():
                if column in ['key', 'Spreadsheet Key*']:
                    if sheet.title != 'Edit Assign Organization':
                        sheet.cell(row=start_row, column=cell_value, value=last_index)

                elif column == 'RowID':
                    if isinstance(cell_value, list):
                        cell_value_list = cell_value
                        for i, value in enumerate(cell_value_list):
                            sheet.cell(row=start_row, column=value, value=1)
                    else:
                        sheet.cell(row=start_row, column=cell_value, value=1)

                elif column == 'Change Job Reason':
                    sheet.cell(row=start_row, column=cell_value, value="SUB_CATEGORY_Chg - Company")

                elif column in ('Country Region', 'Country ISO Code', 'Country ISO3 Code', 'Country ISO2 Code'):
                    if isinstance(cell_value, list):
                        cell_value_list = cell_value
                        for i, value in enumerate(cell_value_list):
                            sheet.cell(row=start_row, column=value, value=row[column])
                    else:
                        sheet.cell(row=start_row, column=cell_value, value=row[column])

                elif column in ('Bank Sort Code'):
                    if isinstance(cell_value, list):
                        cell_value_list = cell_value
                        for i, value in enumerate(cell_value_list):
                            sheet.cell(row=start_row, column=value, value=row[column])


                elif column in ('Currency', 'Compensation Currency'):
                    if isinstance(cell_value, list):
                        cell_value_list = columns_mapping[column]  # Get the list of values for the key
                        for i, cell_value in enumerate(cell_value_list):
                            sheet.cell(row=start_row, column=cell_value, value=currency_code)
                    else:
                        sheet.cell(row=start_row, column=cell_value, value=currency_code)

                else:
                    if isinstance(cell_value, list):
                        cell_value_list = cell_value
                        for i, value in enumerate(cell_value_list):
                            sheet.cell(row=start_row, column=value, value=row[column])
                    else:
                        sheet.cell(row=start_row, column=cell_value, value=row[column])

        if load == 'One Time Payments':
            idx += 1
            sheet.cell(row=start_row, column=8, value=idx)

        # if load == 'NCA Additional Data Position':
        #     start_row += 1

        if load == "Supervisory Org":
            #sheet.cell(row=start_row, column=5, value="N")
            #sheet.cell(row=start_row, column=10, value="N")
            #sheet.cell(row=start_row, column=12, value="Y")
            sheet.cell(row=start_row, column=17, value="SUPERVISORY")
            sheet.cell(row=start_row, column=19, value="Everyone")
            sheet.cell(row=start_row, column=18, value="SUPERVISORY")

            if row['Staffing Model'] == 'Position Management':
                sheet.cell(row=start_row, column=15, value="Y")
            else:
                sheet.cell(row=start_row, column=15, value="N")

        if load == 'Edit Position Restrictions Additional Data':
            idx += 1
            sheet.cell(row=start_row, column=5, value=idx)
        
        if load == 'Company':
            idx += 1
            
            sheet.cell(row=start_row, column=12, value="Y")
            sheet.cell(row=start_row, column=16, value=idx)
            sheet.cell(row=start_row, column=19, value="Org_SubType_Companies")
            
        if load == 'Job Family':
            sheet.cell(row=start_row, column=3, value="Y")
            sheet.cell(row=start_row, column=9, value="N")

        if load == 'Job Family Group':
            idx += 1

            job_family_group_id = row['Job Family Group ID']
            job_family_id = row['Attached Job Family Id']
            job_family_id = str(job_family_group_id) + "_" + str(job_family_id)

            if job_family_group_id not in temp_dict.keys():
                sheet.cell(row=start_row, column=2, value=comp_grad_count)
                temp_dict[job_family_group_id] = comp_grad_count
                comp_grad_count += 1
                comp_profile_count = 1
            else:
                sheet.cell(row=start_row, column=2, value=temp_dict[job_family_group_id])

            if job_family_id not in temp_dict.keys():
                sheet.cell(row=start_row, column=10, value=comp_profile_count)
                temp_dict[job_family_id] = comp_profile_count
                comp_profile_count += 1
            else:
                sheet.cell(row=start_row, column=10, value=temp_dict[job_family_id])

        if load == 'Cost Center':
            #sheet.cell(row=start_row, column=3, value="Y")
            sheet.cell(row=start_row, column=12, value="Y")
            sheet.cell(row=start_row, column=13, value="Y")
            sheet.cell(row=start_row, column=15, value="9c875610c4fc496499e39741b6541dbc")
            sheet.cell(row=start_row, column=20, value="Org_SubType_Cost_Center")
            

        if load == "Location Hierarchy":
            sheet.cell(row=start_row, column=17, value="Location Hierarchy")
            sheet.cell(row=start_row, column=18, value="Location Hierarchy")
            sheet.cell(row=start_row, column=19, value="Everyone")

        if load == "Custom Organizations":
            sheet.cell(row=start_row, column=19, value="Everyone")
            if row['Organization Type'] == "Business Area Hierarchy":
                sheet.cell(row=start_row, column=16, value=row['Superior Organization Name'])
                sheet.cell(row=start_row, column=21, value="")

        if load == "Cost Center Hierarchy":
            sheet.cell(row=start_row, column=19, value="Everyone")

        
        if load == "Job Classification":
            #sheet.cell(row=start_row, column=3, value="Y")
            job_class_group_id = row["Job Classification Group Id"]
            
            if job_class_group_id not in temp_dict.keys():
                sheet.cell(row=start_row, column=2, value=comp_grad_count)
                temp_dict[job_class_group_id] = comp_grad_count
                comp_grad_count += 1
                comp_profile_count = 1
            else:
                sheet.cell(row=start_row, column=2, value=temp_dict[job_class_group_id])

            sheet.cell(row=start_row, column=12, value=comp_profile_count)
            comp_profile_count += 1
        
        if load == "Location":
            idx += 1
            address_row = 1
            sheet.cell(row=start_row, column=3, value="Y")
            sheet.cell(row=start_row, column=44, value=1)
            sheet.cell(row=start_row, column=71, value=idx)
            sheet.cell(row=start_row, column=73, value=idx)
            sheet.cell(row=start_row, column=72, value="Y")
            sheet.cell(row=start_row, column=74, value="Y")
            sheet.cell(row=start_row, column=75, value="BUSINESS")
            sheet.cell(row=start_row, column=77, value="COMMUNICATION_USAGE_BEHAVIOR_TENANTED-6-10")

            if pd.notna(row['Primary Address - Line 1']):
                sheet.cell(row=start_row, column=2, value=last_index)
                sheet.cell(row=start_row, column=53, value=address_row)
                sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_1")
                sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 1'])
                address_row += 1

            if pd.notna(row['Primary Address - Line 2']):
                start_row += 1
                sheet.cell(row=start_row, column=2, value=last_index)
                sheet.cell(row=start_row, column=53, value=address_row)
                sheet.cell(row=start_row, column=44, value=1)
                sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_2")
                sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 2'])
                address_row += 1

            if pd.notna(row['Primary Address - Line 3']):
                start_row += 1
                sheet.cell(row=start_row, column=2, value=last_index)
                sheet.cell(row=start_row, column=53, value=address_row)
                sheet.cell(row=start_row, column=44, value=1)
                sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_3")
                sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 3'])
                address_row += 1

            if pd.notna(row['Primary Address - Line 4']):
                start_row += 1
                sheet.cell(row=start_row, column=2, value=last_index)
                sheet.cell(row=start_row, column=53, value=address_row)
                sheet.cell(row=start_row, column=44, value=1)
                sheet.cell(row=start_row, column=55, value="ADDRESS_LINE_4")
                sheet.cell(row=start_row, column=56, value=row['Primary Address - Line 4'])
                address_row += 1

        if load == "Job Category":
            pass
            # sheet.cell(row=start_row, column=12, value=idx)

        if load == "Job Profile":
            idx += 1
            
            sheet.cell(row=start_row, column=7, value="N")

            if pd.notna(row['Job Family*']):
                sheet.cell(row=start_row, column=19, value=idx)
                sheet.cell(row=start_row, column=21, value=row['Job Family*'])

            if pd.notna(row['Job Classification*']):
                sheet.cell(row=start_row, column=29, value=idx)
                sheet.cell(row=start_row, column=31, value=row['Job Classification*'])

            if pd.notna(row['Pay Rate Type']):
                sheet.cell(row=start_row, column=32, value=idx)
                sheet.cell(row=start_row, column=32, value=row['Pay Rate Type Country'])
                sheet.cell(row=start_row, column=35, value=row['Pay Rate Type'])

            if pd.notna(row['Job Exempt']) and row['Job Exempt']=="Y":
                sheet.cell(row=start_row, column=36, value=idx)
                sheet.cell(row=start_row, column=38, value=row['Job Profile Exempt - Country/Country Region'])
                sheet.cell(row=start_row, column=40, value=row['Job Exempt'])

            if pd.notna(row["Workers' Compensation Code"]):
                sheet.cell(row=start_row, column=41, value=idx)
                sheet.cell(row=start_row, column=42, value=row["Workers' Compensation Code"])

            if pd.notna(row["Responsibilities"]):
                sheet.cell(row=start_row, column=44, value=idx)
                sheet.cell(row=start_row, column=45, value=row["Responsibilities"])

            if pd.notna(row["Work Experience"]):
                sheet.cell(row=start_row, column=48, value=idx)
                sheet.cell(row=start_row, column=49, value=row["Work Experience"])

            if pd.notna(row["Education"]):
                sheet.cell(row=start_row, column=53, value=idx)
                sheet.cell(row=start_row, column=54, value=row["Education"])

            if pd.notna(row["Languages"]):
                sheet.cell(row=start_row, column=58, value=idx)
                sheet.cell(row=start_row, column=59, value=row["Languages"])

            if pd.notna(row["Competencies"]):
                sheet.cell(row=start_row, column=65, value=idx)
                sheet.cell(row=start_row, column=66, value=row["Competencies"])

            if pd.notna(row["Certifications"]):
                sheet.cell(row=start_row, column=70, value=idx)
                sheet.cell(row=start_row, column=72, value=row["Certifications"])

            if pd.notna(row["Training"]):
                sheet.cell(row=start_row, column=80, value=idx)
                sheet.cell(row=start_row, column=81, value=row["Training"])

            if pd.notna(row["Skills on Job Profile"]):
                sheet.cell(row=start_row, column=86, value=idx)
                sheet.cell(row=start_row, column=87, value=row["Skills on Job Profile"])

            if pd.notna(row["Integration Identifier"]):
                sheet.cell(row=start_row, column=92, value=idx)
                sheet.cell(row=start_row, column=93, value=row["Integration Identifier"])

        if load == "Comp Grade and Grade Profile":

            comp_grad_id = row['Compensation Grade ID']
            comp_grad_profile_id = row['Compensation Grade Profile ID']
            comp_step_id = row['Compensation Step Reference ID']
            comp_grad_profile_id = comp_grad_id + "_" + comp_grad_profile_id

            if comp_grad_id not in temp_dict.keys():
                sheet.cell(row=start_row, column=2, value=comp_grad_count)
                temp_dict[comp_grad_id] = comp_grad_count
                comp_grad_count += 1
                comp_profile_count = 1
                comp_step_count = 1
            else:
                sheet.cell(row=start_row, column=2, value=temp_dict[comp_grad_id])
                

            if comp_grad_profile_id not in temp_dict.keys():
                sheet.cell(row=start_row, column=36, value=comp_profile_count)
                temp_dict[comp_grad_profile_id] = comp_profile_count
                comp_profile_count += 1
            else:
                sheet.cell(row=start_row, column=36, value=temp_dict[comp_grad_profile_id])
                

            if pd.notna(row['Compensation Step Reference ID']):
                comp_step_id = comp_grad_id + "_" + comp_step_id
                if comp_step_id not in temp_dict.keys():
                    sheet.cell(row=start_row, column=60, value=comp_step_count)
                    temp_dict[comp_step_id] = comp_step_count
                    comp_step_count += 1
                else:
                    sheet.cell(row=start_row, column=60, value=temp_dict[comp_step_id])

        if load == 'Supplier':
            idx += 1
            sheet.cell(row=start_row, column=3, value='N')
            sheet.cell(row=start_row, column=331, value='1')
            sheet.cell(row=start_row, column=44, value="SUPPLIER_CATEGORY-All_Suppliers")

        if load in ["Employee Compensation", "Overlapping Employee Compensation", "Future Comp Change"]:
            position_id = row['Legacy Worker ID']
            allowance_plan = row['Allowance Plan']
            bonus_plan = row['Bonus Plan']

            if position_id not in temp_dict.keys():
                temp_dict[position_id] = row['spreadsheet_key']
                comp_profile_count = 1
                comp_step_count = 1
                #comp_grad_count += 1

            if pd.notna(row['Hourly Plan']):
                sheet.cell(row=start_row, column=19, value=1)
                sheet.cell(row=start_row, column=20, value=row["Hourly Plan"])
                sheet.cell(row=start_row, column=21, value=row["Hourly Amount"])
                sheet.cell(row=start_row, column=24, value=row["Hourly Currency"])
                sheet.cell(row=start_row, column=25, value=row["Hourly Frequency"])

            if pd.notna(row['Salary Plan']):
                sheet.cell(row=start_row, column=19, value=1)

            sheet.cell(row=start_row, column=2, value=temp_dict[position_id])
            if pd.notna(row['Allowance Plan']):
                allowance_plan = str(position_id) + "_" + allowance_plan
                # if allowance_plan not in temp_dict.keys():
                #     temp_dict[allowance_plan] = comp_profile_count
                    
                #     sheet.cell(row=start_row, column=37, value=temp_dict[allowance_plan])
                # else:
                #     sheet.cell(row=start_row, column=37, value=temp_dict[allowance_plan])
                    
                sheet.cell(row=start_row, column=37, value=row['Allowance Row ID'])
                sheet.cell(row=start_row, column=38, value=row["Allowance Plan"])
                

                #don't populate currency and amount if a plan is default currency or capped plans
                if row["Allowance Plan"] in ['NCA_ALLOWANCE_PERCENT_PLAN-6-3','NCA_ALLOWANCE_PERCENT_PLAN-6-4','NCA_ALLOWANCE_PERCENT_PLAN-6-5','NCA_ALLOWANCE_PERCENT_PLAN-6-6','NCA_ALLOWANCE_PERCENT_PLAN-6-7',
                    'NCA_ALLOWANCE_PERCENT_PLAN-3-20','NCA_ALLOWANCE_PERCENT_PLAN-6-9','NCA_ALLOWANCE_PERCENT_PLAN-6-8','NCA_ALLOWANCE_AMOUNT_PLAN-6-12','NCA_ALLOWANCE_PERCENT_PLAN-3-21',
                    'NCA_ALLOWANCE_PERCENT_PLAN-6-12','NCA_ALLOWANCE_AMOUNT_PLAN-3-35','NCA_ALLOWANCE_PERCENT_PLAN-6-11','NCA_ALLOWANCE_AMOUNT_PLAN-6-14','NCA_ALLOWANCE_AMOUNT_PLAN-6-15']:
                                        
                    sheet.cell(row=start_row, column=39, value="")
                    sheet.cell(row=start_row, column=40, value="")
                    sheet.cell(row=start_row, column=42, value="")
                    sheet.cell(row=start_row, column=43, value="")
                else:
                    sheet.cell(row=start_row, column=40, value=row["Allowance Amount"])
                    sheet.cell(row=start_row, column=39, value=row["Allowance Percentage"])
                    sheet.cell(row=start_row, column=42, value=row["Allowance Currency"])
                    sheet.cell(row=start_row, column=43, value=row["Allowance Frequency"])
                
                if row['Allowance Plan'] in ['NCA_ALLOWANCE_PERCENT_PLAN-6-17', 'NCA_ALLOWANCE_PERCENT_PLAN-6-3', 'NCA_ALLOWANCE_PERCENT_PLAN-6-4', 'NCA_ALLOWANCE_PERCENT_PLAN-6-5', 'NCA_ALLOWANCE_PERCENT_PLAN-6-6', 'NCA_ALLOWANCE_PERCENT_PLAN-6-7', 'NCA_ALLOWANCE_PERCENT_PLAN-3-20', 'NCA_ALLOWANCE_PERCENT_PLAN-6-9', 'NCA_ALLOWANCE_PERCENT_PLAN-6-8', 'NCA_ALLOWANCE_PERCENT_PLAN-6-10', 'NCA_ALLOWANCE_PERCENT_PLAN-3-21', 'NCA_ALLOWANCE_PERCENT_PLAN-6-12',
                        'NCA_ALLOWANCE_PERCENT_PLAN-6-13', 'NCA_ALLOWANCE_PERCENT_PLAN-6-11', 'NCA_ALLOWANCE_PERCENT_PLAN-6-14', 'NCA_ALLOWANCE_PERCENT_PLAN-6-1',
                        'NCA_ALLOWANCE_PERCENT_PLAN-6-18']:
                        sheet.cell(row=start_row, column=42, value="")
                
                if allowance_plan in temp_dict.keys():
                    comp_profile_count += 1

            if pd.notna(row['Bonus Plan']):
                bonus_plan = str(position_id) + "_" + bonus_plan
                # if bonus_plan not in temp_dict.keys():
                #     temp_dict[bonus_plan] = comp_step_count
                #     sheet.cell(row=start_row, column=59, value=row['Bonus Row ID'])
                #     sheet.cell(row=start_row, column=60, value=row["Bonus Plan"])
                #     sheet.cell(row=start_row, column=61, value=row["Bonus Amount"])
                #     sheet.cell(row=start_row, column=62, value=row["Bonus Percentage"])
                #     comp_step_count += 1
                # else:
                #     sheet.cell(row=start_row, column=59, value=row['Bonus Row ID'])
                #     sheet.cell(row=start_row, column=60, value=row["Bonus Plan"])
                #     sheet.cell(row=start_row, column=61, value=row["Bonus Amount"])
                #     sheet.cell(row=start_row, column=62, value=row["Bonus Percentage"])
                sheet.cell(row=start_row, column=59, value=row['Bonus Row ID'])
                sheet.cell(row=start_row, column=60, value=row["Bonus Plan"])
                sheet.cell(row=start_row, column=61, value=row["Bonus Amount"])
                sheet.cell(row=start_row, column=62, value=row["Bonus Percentage"])
                    

        if load == 'Create Position':
            
            if sheet.title == 'Request Default Compensation':
                position_id = row['spreadsheet_key']
                allowance_plan = row['Allowance Plan']
                bonus_plan = row['Bonus Plan']
                
                if position_id not in temp_dict.keys():
                    temp_dict[position_id] = row['spreadsheet_key']
                    comp_profile_count = 1
                    comp_step_count = 1

                if pd.notna(row['Hourly Plan']):
                    sheet.cell(row=start_row, column=11, value=1)
                    sheet.cell(row=start_row, column=12, value=row["Hourly Plan"])
                    sheet.cell(row=start_row, column=13, value=row["Hourly Amount"])
                    sheet.cell(row=start_row, column=16, value=row["Hourly Currency"])
                    sheet.cell(row=start_row, column=17, value=row["Hourly Frequency"])

                if pd.notna(row['Salary Plan']):
                    sheet.cell(row=start_row, column=11, value=1)

                sheet.cell(row=start_row, column=2, value=temp_dict[position_id])
                
                if pd.notna(row['Allowance Plan']):

                    allowance_plan = str(position_id) + "_" + allowance_plan

                    if allowance_plan not in temp_dict.keys():
                        temp_dict[allowance_plan] = comp_profile_count
                        #comp_profile_count += 1
                        sheet.cell(row=start_row, column=29, value=temp_dict[allowance_plan])
                    else:
                        sheet.cell(row=start_row, column=29, value=temp_dict[allowance_plan])

                    sheet.cell(row=start_row, column=30, value=row["Allowance Plan"])

                    #don't populate currency and amount if a plan is default currency or capped plans
                    if row["Allowance Plan"] in ['NCA_ALLOWANCE_PERCENT_PLAN-6-3','NCA_ALLOWANCE_PERCENT_PLAN-6-4','NCA_ALLOWANCE_PERCENT_PLAN-6-5','NCA_ALLOWANCE_PERCENT_PLAN-6-6','NCA_ALLOWANCE_PERCENT_PLAN-6-7',
                        'NCA_ALLOWANCE_PERCENT_PLAN-3-20','NCA_ALLOWANCE_PERCENT_PLAN-6-9','NCA_ALLOWANCE_PERCENT_PLAN-6-8','NCA_ALLOWANCE_AMOUNT_PLAN-6-12','NCA_ALLOWANCE_PERCENT_PLAN-3-21',
                        'NCA_ALLOWANCE_PERCENT_PLAN-6-12','NCA_ALLOWANCE_AMOUNT_PLAN-3-35','NCA_ALLOWANCE_PERCENT_PLAN-6-11','NCA_ALLOWANCE_AMOUNT_PLAN-6-14','NCA_ALLOWANCE_AMOUNT_PLAN-6-15']:
            
                        sheet.cell(row=start_row, column=31, value="")
                        sheet.cell(row=start_row, column=32, value="")
                        sheet.cell(row=start_row, column=34, value="")
                        sheet.cell(row=start_row, column=35, value="")
                    else:
                        sheet.cell(row=start_row, column=31, value=row["Allowance Percentage"])
                        sheet.cell(row=start_row, column=32, value=row["Allowance Amount"])
                        sheet.cell(row=start_row, column=34, value=row["Allowance Currency"])
                        sheet.cell(row=start_row, column=35, value=row["Allowance Frequency"])

                    #these are the plans that should not have any currency
                    if row['Allowance Plan'] in ['NCA_ALLOWANCE_PERCENT_PLAN-6-17', 'NCA_ALLOWANCE_PERCENT_PLAN-6-3', 'NCA_ALLOWANCE_PERCENT_PLAN-6-4', 'NCA_ALLOWANCE_PERCENT_PLAN-6-5', 'NCA_ALLOWANCE_PERCENT_PLAN-6-6', 'NCA_ALLOWANCE_PERCENT_PLAN-6-7', 'NCA_ALLOWANCE_PERCENT_PLAN-3-20', 'NCA_ALLOWANCE_PERCENT_PLAN-6-9', 'NCA_ALLOWANCE_PERCENT_PLAN-6-8', 'NCA_ALLOWANCE_PERCENT_PLAN-6-10', 'NCA_ALLOWANCE_PERCENT_PLAN-3-21', 'NCA_ALLOWANCE_PERCENT_PLAN-6-12',
                        'NCA_ALLOWANCE_PERCENT_PLAN-6-13', 'NCA_ALLOWANCE_PERCENT_PLAN-6-11', 'NCA_ALLOWANCE_PERCENT_PLAN-6-14', 'NCA_ALLOWANCE_PERCENT_PLAN-6-1',
                        'NCA_ALLOWANCE_PERCENT_PLAN-6-18']:
                        sheet.cell(row=start_row, column=34, value="")
                    if allowance_plan in temp_dict.keys():
                        comp_profile_count += 1

                if pd.notna(row['Bonus Plan']):
                    bonus_plan = str(position_id) + "_" + bonus_plan
                    if bonus_plan not in temp_dict.keys():
                        temp_dict[bonus_plan] = comp_step_count
                        
                        sheet.cell(row=start_row, column=51, value=temp_dict[bonus_plan])
                        sheet.cell(row=start_row, column=52, value=row["Bonus Plan"])
                        sheet.cell(row=start_row, column=53, value=row["Bonus Amount"])
                        sheet.cell(row=start_row, column=54, value=row["Bonus Percentage"])
                        comp_step_count += 1
                    else:
                        sheet.cell(row=start_row, column=51, value=temp_dict[bonus_plan])
                        
                        sheet.cell(row=start_row, column=52, value=row["Bonus Plan"])
                        sheet.cell(row=start_row, column=53, value=row["Bonus Amount"])
                        sheet.cell(row=start_row, column=54, value=row["Bonus Percentage"])

                if pd.notna(row['Hourly Plan']) or pd.notna(row['Salary Plan']) or pd.notna(row['Allowance Plan']) or pd.notna(row['Bonus Plan']):
                    sheet.cell(row=start_row, column=2, value=temp_dict[position_id])
                else:
                    sheet.cell(row=start_row, column=2, value="")


        if load == 'Put Candidate':
  
            sheet.cell(row=start_row, column=3, value='Y')

            sheet.cell(row=start_row, column=67, value='')
           
            #start_row += 1

        if load in ['Prehire', "Future Prehire"]:
            idx += 1
            #toogle for update
            sheet.cell(row=start_row, column=3, value='Y')

            if pd.isna(row['Email Address']):
                pass
            else:
                sheet.cell(row=start_row, column=353, value=row['Row_Count'])
                sheet.cell(row=start_row, column=358, value=1)
                sheet.cell(row=start_row, column=360, value=1)
            
                #default if visibility as Y and primary email as Y when its blank
                if pd.isna(row['Visibility (Email)']):
                    sheet.cell(row=start_row, column=359, value='Y')

                if pd.isna(row['Primary (Email)']):
                    sheet.cell(row=start_row, column=361, value='Y')

                if pd.isna(row['Usage Type (Email)']):
                    sheet.cell(row=start_row, column=362, value='WORK')
            
            if pd.isnull(row['Preferred First Name']) or row['Preferred First Name'].strip() == '':
                sheet.cell(row=start_row, column=39, value=" ")
            # if pd.isnull(row['Preferred First Name']) or row['Preferred First Name'].strip() == '':
            #     sheet.cell(row=start_row, column=43, value=row['Legal First Name'])
            #     sheet.cell(row=start_row, column=44, value=row['Legal Middle Name'])
            #     sheet.cell(row=start_row, column=45, value=row['Legal Last Name'])


        if load in ['Hire Employee', 'Overlapping Hire Employee', 'Future Hires']:
            
            #hardcode Hire reason to Conversion
            if sheet.title == 'Hire Employee':
                sheet.cell(row=start_row, column=410, value = "HIRE_EMPLOYEE-Conversion")

        if load in ['Hire CWR', 'Future Hire CWR']:
            
            #hardcode Hire reason to Conversion
            if sheet.title == 'Contract Contingent Worker':
                sheet.cell(row=start_row, column=409, value = "CONTRACT_CONTINGENT_WORKER-Conversion")
                sheet.cell(row=start_row, column=447, value = "1")

            # if sheet.title == 'Edit Assign Organization':
            #     sheet.cell(row=start_row, column=2, value=row['Spreadsheet Key*'])


        if load == 'Legacy ID':
            sheet.cell(row=start_row, column=5, value="N")
            sheet.cell(row=start_row, column=6, value="1")
            sheet.cell(row=start_row, column=10, value="LEGACY_ID")

        if load == 'Payment Elections':
            sheet.cell(row=start_row, column=10, value="1")
            sheet.cell(row=start_row, column=17, value="Payment_Type_DirectDep")
            sheet.cell(row=start_row, column=11, value="PAYMENT_ELECTION_RULE-ExtPaySys")
            sheet.cell(row=start_row, column=14, value="PAYMENT_ELECTION_RULE-ExtPaySys")

        if load == "Passports":
            sheet.cell(row=start_row, column=6, value="1")
            sheet.cell(row=start_row, column=7, value="N")
            sheet.cell(row=start_row, column=5, value="N")

        if load == "Visas":
            sheet.cell(row=start_row, column=17, value="1")
            sheet.cell(row=start_row, column=18, value="N")
            sheet.cell(row=start_row, column=5, value="N")

        if load in ["Personal Contact Change"]:
            
            if pd.isna(row['Address ID']) or row['Address ID']=="":
                sheet.cell(row=start_row, column=5, value="")
                sheet.cell(row=start_row, column=12, value="")
                sheet.cell(row=start_row, column=14, value="")
                sheet.cell(row=start_row, column=21, value="")
                sheet.cell(row=start_row, column=32, value="")
                sheet.cell(row=start_row, column=33, value="")
                sheet.cell(row=start_row, column=34, value="")
                sheet.cell(row=start_row, column=35, value="")
                sheet.cell(row=start_row, column=36, value="")

            if pd.notna(row['Home Address Line 1']):
                sheet.cell(row=start_row, column=5, value=row['Address Row ID'])
                sheet.cell(row=start_row, column=14, value="1")
                sheet.cell(row=start_row, column=16, value="ADDRESS_LINE_1")
                sheet.cell(row=start_row, column=20, value=row['Country'])
                sheet.cell(row=start_row, column=32, value="1")
                sheet.cell(row=start_row, column=33, value=row["Public"])
                sheet.cell(row=start_row, column=34, value="1")
                sheet.cell(row=start_row, column=35, value=row["Primary"])
                sheet.cell(row=start_row, column=36, value="HOME")
                #sheet.cell(row=start_row, column=38, value="COMMUNICATION_USAGE_BEHAVIOR_TENANTED-6-13")

            if pd.notna(row['Home Address Line 1 - Local']):
                sheet.cell(row=start_row, column=5, value=row['Address Row ID'])
                sheet.cell(row=start_row, column=14, value="1")
                sheet.cell(row=start_row, column=16, value="ADDRESS_LINE_1_LOCAL")
                sheet.cell(row=start_row, column=17, value=row['Home Address Line 1 - Local'])
                sheet.cell(row=start_row, column=20, value=row['Country'])
                sheet.cell(row=start_row, column=32, value="1")
                sheet.cell(row=start_row, column=33, value=row["Public"])
                sheet.cell(row=start_row, column=34, value="1")
                sheet.cell(row=start_row, column=35, value=row["Primary"])
                sheet.cell(row=start_row, column=36, value="HOME")

            
            if pd.notna(row['Home Phone Number']) and row['Home Phone Number'] !="":
                #sheet.cell(row=start_row, column=5, value="1")
                sheet.cell(row=start_row, column=44, value=row['Home Phone Row ID'])
                sheet.cell(row=start_row, column=59, value="1")
                sheet.cell(row=start_row, column=61, value="1")
                sheet.cell(row=start_row, column=60, value=row['Phone Visibility'])

            if pd.notna(row['Home Email']) and row['Home Email'] !="":
                sheet.cell(row=start_row, column=69, value=row['Home Email Row ID'])
                sheet.cell(row=start_row, column=70, value="N")
                sheet.cell(row=start_row, column=74, value="1")
                sheet.cell(row=start_row, column=75, value=row["Email Visibility"])
                sheet.cell(row=start_row, column=76, value="1")
                sheet.cell(row=start_row, column=77, value="Y")
                sheet.cell(row=start_row, column=78, value=row["Email Usage"])

            if pd.notna(row['Home Address Line 2']) and row['Number of Address']==1:
                # Write address line 2 to the new row
                start_row += 1
                sheet.cell(row=start_row, column=2, value=row['spreadsheet_key'])
                sheet.cell(row=start_row, column=3, value=row['Legacy Worker ID'])

            # Check if address line 2 is not blank or na
            if pd.notna(row['Home Address Line 2']):
                # write additional n for home address line 2
                sheet.cell(row=start_row, column=5, value=row['Address Row ID'])
                sheet.cell(row=start_row, column=14, value="2")
                sheet.cell(row=start_row, column=16, value="ADDRESS_LINE_2")
                sheet.cell(row=start_row, column=17, value=row['Home Address Line 2'])
                if pd.notna(row['Home Usage Type']):
                    sheet.cell(row=start_row, column=14, value="2")

            # Check if address line 3 is not blank, then create a new row and add following n
            if pd.notna(row['Home Address Line 3']) and row['Number of Address']==1:
                start_row += 1
                # Write address line 3 to the new row
                sheet.cell(row=start_row, column=2, value=row['spreadsheet_key'])
                sheet.cell(row=start_row, column=3, value=row['Legacy Worker ID'])

            if pd.notna(row['Home Address Line 3']):    
                sheet.cell(row=start_row, column=5, value=row['Address Row ID'])
                sheet.cell(row=start_row, column=14, value="3")
                sheet.cell(row=start_row, column=16, value="ADDRESS_LINE_3")
                sheet.cell(row=start_row, column=17, value=row['Home Address Line 3'])
                if pd.notna(row['Home Usage Type']):
                    sheet.cell(row=start_row, column=14, value="3")
            
            # Check if address line 4 is not blank, then create a new row and add following n
            if pd.notna(row['Home Address Line 4']) and row['Number of Address']==1:
                start_row += 1
                # Write address line 4 to the new row
                sheet.cell(row=start_row, column=2, value=row['spreadsheet_key'])
                sheet.cell(row=start_row, column=3, value=row['Legacy Worker ID'])

            if pd.notna(row['Home Address Line 4']):    
                sheet.cell(row=start_row, column=5, value=row['Address Row ID'])
                sheet.cell(row=start_row, column=14, value="3")
                sheet.cell(row=start_row, column=16, value="ADDRESS_LINE_4")
                sheet.cell(row=start_row, column=17, value=row['Home Address Line 4'])
                if pd.notna(row['Home Usage Type']):
                    sheet.cell(row=start_row, column=14, value="4")

            if pd.notna(row['Home Address Line 2 - Local']) and row['Number of Address']==1:
                # Write address line 2 to the new row
                start_row += 1
                sheet.cell(row=start_row, column=2, value=row['spreadsheet_key'])
                sheet.cell(row=start_row, column=3, value=row['Legacy Worker ID'])
                
             # Check if address line 2 is not blank or na
            if pd.notna(row['Home Address Line 2 - Local']):
                # write additional n for home address line 2
                sheet.cell(row=start_row, column=5, value=row['Address Row ID'])
                sheet.cell(row=start_row, column=14, value="2")
                sheet.cell(row=start_row, column=16, value="ADDRESS_LINE_2_LOCAL")
                sheet.cell(row=start_row, column=17, value=row['Home Address Line 2 - Local'])
                if pd.notna(row['Home Usage Type']):
                    sheet.cell(row=start_row, column=14, value="2")

            # Check if address line 3 is not blank, then create a new row and add following n
            if pd.notna(row['Home Address Line 3 - Local']) and row['Number of Address']==1:
                start_row += 1
                # Write address line 3 to the new row
                sheet.cell(row=start_row, column=2, value=row['spreadsheet_key'])
                sheet.cell(row=start_row, column=3, value=row['Legacy Worker ID'])

            if pd.notna(row['Home Address Line 3 - Local']):
                sheet.cell(row=start_row, column=5, value=row['Address Row ID'])
                sheet.cell(row=start_row, column=14, value="3")
                sheet.cell(row=start_row, column=16, value="ADDRESS_LINE_3_LOCAL")
                sheet.cell(row=start_row, column=17, value=row['Home Address Line 3 - Local'])
                if pd.notna(row['Home Usage Type']):
                    sheet.cell(row=start_row, column=14, value="3")
                    

        if load in ["Work Contact Change"]:
            if pd.isna(row['Work Phone Number']):
                sheet.cell(row=start_row, column=44, value="")
                sheet.cell(row=start_row, column=60, value="")
               
            if pd.notna(row['Email Address']):
                sheet.cell(row=start_row, column=69, value=row["Email Row ID"])
                sheet.cell(row=start_row, column=70, value="N")
                sheet.cell(row=start_row, column=74, value="1")
                sheet.cell(row=start_row, column=75, value=row["Email Public"])
                sheet.cell(row=start_row, column=76, value="1")
                sheet.cell(row=start_row, column=77, value=row["Email Primary"])
                sheet.cell(row=start_row, column=78, value="WORK")

        if load == "Emergency Contact":

            if pd.notna(row['Home Address Line 1']):
                sheet.cell(row=start_row, column=107, value="1")
                sheet.cell(row=start_row, column=114, value=row['Country'])
                sheet.cell(row=start_row, column=116, value="1")
                sheet.cell(row=start_row, column=118, value="ADDRESS_LINE_1")
                sheet.cell(row=start_row, column=119, value=row['Home Address Line 1'])
                sheet.cell(row=start_row, column=122, value=row['Country'])
                sheet.cell(row=start_row, column=134, value="1")
                sheet.cell(row=start_row, column=135, value="N")
                sheet.cell(row=start_row, column=136, value="1")
                sheet.cell(row=start_row, column=137, value="Y")
                sheet.cell(row=start_row, column=138, value="HOME")
                sheet.cell(row=start_row, column=140, value="COMMUNICATION_USAGE_BEHAVIOR_TENANTED-6-13")

            if pd.notna(row['Home Email']):
                sheet.cell(row=start_row, column=171, value="1")
                sheet.cell(row=start_row, column=172, value="N")
                sheet.cell(row=start_row, column=176, value="1")
                sheet.cell(row=start_row, column=177, value="N")
                sheet.cell(row=start_row, column=178, value="1")
                sheet.cell(row=start_row, column=179, value="Y")
                sheet.cell(row=start_row, column=180, value="HOME")

            if pd.notna(row['Home Address Line 2']):
                if row['Number of Phone']<2:
                    start_row += 1
                
                # Write address line 3 to the new row
                sheet.cell(row=start_row, column=2, value=row['spreadsheet_key'])
                sheet.cell(row=start_row, column=3, value=row['Legacy Worker ID'])
                sheet.cell(row=start_row, column=5, value=row['Contact Row ID'])
                sheet.cell(row=start_row, column=106, value=row['Contact Row ID'])
                sheet.cell(row=start_row, column=107, value="1")
                sheet.cell(row=start_row, column=114, value=row['Country'])
                sheet.cell(row=start_row, column=116, value="2")
                sheet.cell(row=start_row, column=118, value="ADDRESS_LINE_2")
                sheet.cell(row=start_row, column=119, value=row['Home Address Line 2'])

            if pd.isnull(row['Phone Numbers']) or row['Phone Numbers']=="":
                sheet.cell(row=start_row, column=146, value="")
                sheet.cell(row=start_row, column=156, value="")
                sheet.cell(row=start_row, column=161, value="")
                sheet.cell(row=start_row, column=162, value="")
                sheet.cell(row=start_row, column=163, value="")

            #start_row += 1
            last_worker = row["Legacy Worker ID"]
        start_row += 1

    return sheet, all_unique_data_list, unavailable_reference_id_list


def load_file_params(load, base_path):

    # Define the dictionary with load conditions as keys and corresponding values
    params = {
        "Change Personal Information": {
            "input_sheet": "Personal-Info",
            "eib_file_name": base_path + "EIB Template\\Change_Personal_Information_v41.2.xlsx"
        },
        "Pronoun Public Preference": {
            "input_sheet": "Personal-Info",
            "eib_file_name": base_path + "EIB Template\\Change_Public_Profile_Preferences_v42.1.xlsx"
        },
        "Change Legal Name": {
            "input_sheet": "Basic Info - Names",
            "eib_file_name": base_path + "EIB Template\\Change_Legal_Name_v40.1.xlsx"
        },
        "Change Preferred Name": {
            "input_sheet": "Basic Info - Names",
            "eib_file_name": base_path + "EIB Template\\Change_Preferred_Name_v40.1.xlsx"
        },
        "Change Job": {
            "input_sheet": "Basic Job Info",
            "eib_file_name": base_path + "EIB Template\\Change_Job_v40.2.xlsx"
        },
        "Change Job Assign Pay Group": {
            "input_sheet": "Basic Job Info",
            "eib_file_name": base_path + "EIB Template\\Change_Job_v40.2.xlsx"
        },
        "Change Job Change Org": {
            "input_sheet": "Basic Job Info",
            "eib_file_name": base_path + "EIB Template\\Change_Job_v40.1.xlsx"
        },
        "Legacy ID": {
            "input_sheet": "Basic Job Info",
            "eib_file_name": base_path + "EIB Template\\Change_Other_IDs_v40.1.xlsx"
        },
        "Service Dates": {
            "input_sheet": "Hire-Employee",
            "eib_file_name": base_path + "EIB Template\\Edit_Service_Dates_v42.0.xlsx"
        },
        "Contingent Service Dates": {
            "input_sheet": "Hire-Contingent-Worker",
            "eib_file_name": base_path + "EIB Template\\Edit_Service_Dates_v42.0.xlsx"
        },
        "National ID": {
            "input_sheet": "National-ID-Info",
            "eib_file_name": base_path + "EIB Template\\Change_Government_IDs_v41.2.xlsx"
        },
        "Visas": {
            "input_sheet": "Visa-Info",
            "eib_file_name": base_path + "EIB Template\\Change_Visas_v41.2.xlsx"
        },
        "Passports": {
            "input_sheet": "Passport-Info",
            "eib_file_name": base_path + "EIB Template\\Change_Passports_v41.2.xlsx"
        },
        "Payment Elections": {
            "input_sheet": "Banking-Details",
            "eib_file_name": base_path + "EIB Template\\Submit_Payment_Election_Enrollment_v41.2.xlsx"
        },
        "Personal Contact Change": {
            "input_sheet": "Worker-Personal-Contact-Data",
            "eib_file_name": base_path + "EIB Template\\Maintain_Personal_Contact_Information_v40.2.xlsx"
        },
        "Work Contact Change": {
            "input_sheet": "Worker-Work-Contact-Data",
            "eib_file_name": base_path + "EIB Template\\Maintain_Work_Contact_Information_v40.2.xlsx"
        },
        "Absence Input": {
            "input_sheet": "Time-Off-Plan-Override",
            "eib_file_name": base_path + "EIB Template\\Put_Override_Balance_v41.1.xlsx"
        },
        "Emergency Contact": {
            "input_sheet": "Emergency-Contact",
            "eib_file_name": base_path + "EIB Template\\Change_Emergency_Contacts_v41.2.xlsx"
        }, 
        "Supervisory Org": {
            "input_sheet": "Supervisory-Org",
            "eib_file_name": base_path + "EIB Template\\Add_Update_Supervisory_Organization_v41.2.xlsx"
        },
        "Job Requisitions": {
            "input_sheet": "Job-Requisitions",
            "eib_file_name": base_path + "EIB Template\\Create_Job_Requisition_v41.2.xlsx"
        },
        "Job Requisition Roles": {
            "input_sheet": "Job-Requisition-Roles",
            "eib_file_name": base_path + "EIB Template\\Job_Req_Assign_Roles_v41.2.xlsx"
        },
        "Create Position": {
            "input_sheet": "Positions",
            "eib_file_name": base_path + "EIB Template\\Create_Position_v41.2.xlsx"
        },
        "Prehire": {
            "input_sheet": "Prehire",
            "eib_file_name": base_path + "EIB Template\\Put_Applicant_v41.2.xlsx"
        },
        "Hire Employee": {
            "input_sheet": "Hire-Employee",
            "eib_file_name": base_path + "EIB Template\\Hire_Employee_v41.2.xlsx"
        },
        "Hire CWR": {
            "input_sheet": "Hire-Contingent-Worker",
            "eib_file_name": base_path + "EIB Template\\Contract_Contingent_Worker_v41.2.xlsx"
        },
        "Change-Job": {
            "input_sheet": "Change-Job",
            "eib_file_name": base_path + "EIB Template\\Change_Job_v41.2.xlsx"
        },
        "Terminate Employee": {
            "input_sheet": "Terminate-Employee",
            "eib_file_name": base_path + "EIB Template\\Terminate_Employee_v41.2.xlsx"
        },
        "End Contingent Worker Contracts": {
            "input_sheet": "End-Contingent-Worker-Contracts",
            "eib_file_name": base_path + "EIB Template\\End_Contingent_Worker_Contract_v41.2.xlsx"
        },
        "One Time Payments": {
            "input_sheet": "One-Time-Payments",
            "eib_file_name": base_path + "EIB Template\\Request_One-Time_Payment_v41.2.xlsx"
        },
        "Compensation": {
            "input_sheet": "Compensation",
            "eib_file_name": base_path + "EIB Template\\Request_Compensation_Change_v41.2.xlsx"
        },
        "Put Candidate": {
            "input_sheet": "Put-Candidate",
            "eib_file_name": base_path + "EIB Template\\Put_Candidate_v41.2.xlsx"
        },
        "Edit Job Requisitions": {
            "input_sheet": "Job-Requisition-Roles",
            "eib_file_name": base_path + "EIB Template\\Edit_Job_Requisition_v41.2.xlsx"
        },
        "Probation Info": {
            "input_sheet": "Employee-Probation-Period-Event",
            "eib_file_name": base_path + "EIB Template\\Manage_Employee_Probation_Periods_Event_v41.2.xlsx"
        },
        "Manage Goals": {
            "input_sheet": "Manage-Goals",
            "eib_file_name": base_path + "EIB Template\\Manage_Goals_v41.2.xlsx"
        },
        "Performance Reviews": {
            "input_sheet": "Performance-Reviews",
            "eib_file_name": base_path + "EIB Template\\Start_Performance_Review_v41.2.xlsx"
        },
        "Supplier": {
            "input_sheet": "Supplier",
            "eib_file_name": base_path + "EIB Template\\Submit_Supplier_v42.0.xlsx"
        },
        "Assign Work Schedule": {
            "input_sheet": "Assign-Work-Schedule",
            "eib_file_name": base_path + "EIB Template\\Assign_Work_Schedule_v42.0.xlsx"
        },
        "Leave Of Absence Events": {
            "input_sheet": "Employee-Leave-Of-Absence",
            "eib_file_name": base_path + "EIB Template\\Request_Leave_of_Absence_v41.2.xlsx"
        },
        "Role Based Assignments": {
            "input_sheet": "Role-Based-Assignments",
            "eib_file_name": base_path + "EIB Template\\Assign_Roles_v41.2.xlsx"
        },
        "Worker Additional Data": {
            "input_sheet": "Worker-Additional-Data",
            "eib_file_name": base_path + "EIB Template\\Edit_Worker_Additional_Data_v42.0.xlsx"
        },
        "NCA Additional Data Position": {
            "input_sheet": "NCA-Additional-Data-Position-No",
            "eib_file_name": base_path + "EIB Template\\Edit_Position_Restrictions_v41.2.xlsx"
        },
        "Edit Position Restrictions Additional Data": {
            "input_sheet": "Edit-Position-Restrictions-Addi",
            "eib_file_name": base_path + "EIB Template"
                                         "\\Edit_Position_Restrictions_Additional_Data_Effective_Dated_EIB_v41.1.xlsx"
        },
        "International Assignments": {
            "input_sheet": "International-Assignments",
            "eib_file_name": base_path + "EIB Template\\Add_Additional_Job_EIB_v41.1.xlsx"
        },
        "Company": {
            "input_sheet": "Company",
            "eib_file_name": base_path + "EIB Template\\Put_Company_Organization_v41.2.xlsx"
        },
        "Job Family": {
            "input_sheet": "Job-Family",
            "eib_file_name": base_path + "EIB Template"
                                         "\\Put_Job_Family_v42.0.xlsx"
        },
        "Job Family Group": {
            "input_sheet": "Job-Family-Group",
            "eib_file_name": base_path + "EIB Template\\Put_Job_Family_Group_v42.0.xlsx"
        },
        "Cost Center": {
            "input_sheet": "Cost-Center",
            "eib_file_name": base_path + "EIB Template\\Put_Cost_Center_v42.0.xlsx"
        },
        "Location Hierarchy": {
            "input_sheet": "Location-Hierarchy",
            "eib_file_name": base_path + "EIB Template\\Add_Update_Location_Hierarchy_Organization_v41.2.xlsx"
        },
        "Custom Organizations": {
            "input_sheet": "Custom-Organization",
            "eib_file_name": base_path + "EIB Template\\Add_Update_Custom_Organization_v41.2.xlsx"
        },
        "Cost Center Hierarchy": {
            "input_sheet": "Cost-Center-Hierarchy",
            "eib_file_name": base_path + "EIB Template\\Add_Update_Costcenter_Hierarchy_Organization_v41.2.xlsx"
        },
        "Collective Agreement": {
            "input_sheet": "Collective-Agreement-Type",
            "eib_file_name": base_path + "EIB Template\\Put_Collective_Agreement_v42.0.xlsx"
        },
        "Job Classification": {
            "input_sheet": "Job-Classification",
            "eib_file_name": base_path + "EIB Template\\Put_Job_Classification_Group_v42.0.xlsx"
        },
        "Put Supervisory Assignment Restrictions": {
            "input_sheet": "Put-Supervisory-Assignment",
            "eib_file_name": base_path + "EIB Template\\Put_Supervisory_Organization_Assignment_Restrictions_v41.2.xlsx"
        },
        "Location": {
            "input_sheet": "Location",
            "eib_file_name": base_path + "EIB Template\\Put_Location_v42.0.xlsx"
        },
        "Job Category": {
            "input_sheet": "Job-Category",
            "eib_file_name": base_path + "EIB Template\\Put_Job_Category_v42.0.xlsx"
        },
        "Job Profile": {
            "input_sheet": "Job-Profile",
            "eib_file_name": base_path + "EIB Template\\Put_Job_Profile_v41.2.xlsx"
        },
        "Comp Grade and Grade Profile": {
            "input_sheet": "Comp-Grade-and-Grade-Profiles",
            "eib_file_name": base_path + "EIB Template\\Put_Compensation_Grade_v41.2.xlsx"
        },
        "Job Requisition Additional Data": {
            "input_sheet": "Job-Requisition-Additional-Data",
            "eib_file_name": base_path + "EIB Template\\Edit_Job_Requisition_Additional_Data_v41.1.xlsx"
        },
        "Employee Compensation": {
            "input_sheet": "Employee-Compensation",
            "eib_file_name": base_path + "EIB Template\\Request_Compensation_Change_v41.2.xlsx"
        },
        "Update Workday Account": {
            "input_sheet": "Update-Workday-Account",
            "eib_file_name": base_path + "EIB Template\\Update_Workday_Account_v41.2.xlsx"
        },
        "Add Workday Account": {
            "input_sheet": "Add-Workday-Account",
            "eib_file_name": base_path + "EIB Template\\Add_Workday_Account_v42.0.xlsx"
        },
        "Overlapping Hire Employee": {
            "input_sheet": "Overlapping-Hire-Employee",
            "eib_file_name": base_path + "EIB Template\\Overlapping_Hire_Employee_v41.2.xlsx"
        },
        "Overlapping Employee Compensation": {
            "input_sheet": "Overlapping-Employee-Comp",
            "eib_file_name": base_path + "EIB Template\\Overlapping_Request_Compensation_Change_v41.2.xlsx"
        },
        "Time Off Events": {
            "input_sheet": "Time-Off-Events",
            "eib_file_name": base_path + "EIB Template\\Enter_Time_Off_v41.2.xlsx"
        },
        "Assign Notice Period": {
            "input_sheet": "Assign-Notice-Period",
            "eib_file_name": base_path + "EIB Template\\Put_Edit_Notice_Periods_Event_v42.0.xlsx"
        },
        "User Based Assignments": {
            "input_sheet": "User-Based-Assignments",
            "eib_file_name": base_path + "EIB Template\\User_2_Assign_Roles_v41.2.xlsx"
        },
        "Licenses": {
            "input_sheet": "Licenses",
            "eib_file_name": base_path + "EIB Template\\Change_Licenses_v41.2.xlsx"
        },
        "Job History":{
            "input_sheet": "Job-History",
            "eib_file_name": base_path + "EIB Template\\Manage_Job_History_v42.0.xlsx"
        },
        "Education": {
            "input_sheet": "Education",
            "eib_file_name": base_path + "EIB Template\\Manage_Education_v42.0.xlsx"
        },
        "Skills": {
            "input_sheet": "Skills",
            "eib_file_name": base_path + "EIB Template\\Manage_Skills_v42.0.xlsx"
        },
        "Other Ids": {
            "input_sheet": "Other-Ids",
            "eib_file_name": base_path + "EIB Template\\Change_Other_IDs_v40.2.xlsx"
        },
        'Flexible Work Arrangements': {
            "input_sheet": "Flexible-Work-Arrangement",
            "eib_file_name": base_path + "EIB Template\\Add_Flexible_Work_Arrangement_v42.0.xlsx"
        },
        "Worker Collective Agreement":  {
            "input_sheet": "Worker-Collective-Agreements",
            "eib_file_name": base_path + "EIB Template\\Assign_Employee_Collective_Agreement_Event_v42.0.xlsx"
        },
        "Job History Company":  {
            "input_sheet": "Job-History-Companies",
            "eib_file_name": base_path + "EIB Template\\Put_Job_History_Company_v42.0.xlsx"
        },
        "Future Termination":  {
            "input_sheet": "Future-Terminations",
            "eib_file_name": base_path + "EIB Template\\Future_Terminate_Employee_v41.2.xlsx"
        },
        "Future Hires":  {
            "input_sheet": "Future-Hire-Employee",
            "eib_file_name": base_path + "EIB Template\\Future_Hire_Employee_v41.2.xlsx"
        },
        "Future Job Change":  {
            "input_sheet": "Job-History-Companies",
            "eib_file_name": base_path + "EIB Template\\.xlsx"
        },
        "Future Comp Change":{
            "input_sheet": "Future-Employee-Compensation",
            "eib_file_name": base_path + "EIB Template\\Future_Request_Compensation_Change_v41.2.xlsx"
        },
        "Skills Reference Data":{
            "input_sheet": "Skills-Reference-Data",
            "eib_file_name": base_path + "EIB Template\\Put_Skill_v42.2.xlsx"
        },
        "Future Prehire": {
            "input_sheet": "Future-Prehire",
            "eib_file_name": base_path + "EIB Template\\Future_Put_Applicant_v41.2.xlsx"
        },
        "Future Hire CWR": {
            "input_sheet": "Future-Hire-Contingent-Worker",
            "eib_file_name": base_path + "EIB Template\\Future_Contract_Contingent_Worker_v41.2.xlsx"
        },
        "End CWR Contract": {
            "input_sheet": "End-CWR-Contract",
            "eib_file_name": base_path + "EIB Template\\Future_End_Contingent_Worker_Contract_v41.2.xlsx"
        },
    }

    # Retrieve the parameters from the dictionary based on theload condition
    params_for_load = params.get(load, {})

    # Extract the input_sheet and eib_file_name values
    input_sheet = params_for_load.get("input_sheet", "")
    eib_file_name = params_for_load.get("eib_file_name", "")

    # Return the values
    return input_sheet, eib_file_name


def get_mapping_data_dict(mapping_file):
    mapped_data = pd.read_excel(mapping_file)
    mapping_data_dict = {}
    for idx, row in mapped_data.iterrows():
        value_dict = {row['Reference ID Type']: row['Reference ID Value']}
        mapping_data_dict[row['Business Object Instance']] = value_dict

    return mapping_data_dict

def update_spreadsheet_key(eib_file_name, sheet_name, input_data):
    
    eib_data = pd.read_excel(eib_file_name, sheet_name=sheet_name, skiprows=[0, 1, 2, 3], dtype=object)
    
    eib_data = eib_data[['Spreadsheet Key*', 'Job Requisition ID']]
    eib_data = eib_data.drop_duplicates().reset_index(drop=True)
    merged_df = pd.merge(input_data, eib_data, on='Job Requisition ID', how='left')
    # Create new column 'spreadsheet_key' in df1 based on matched values in df2
    input_data['spreadsheet_key'] = merged_df['Spreadsheet Key*']
    return input_data

def process_load(load, all_unique_data_list, unavailable_reference_id_list, mapping_file,input_data_temp_file,input_sheet=None):

    # mapping_file = base_path + 'Combined_Mapping.xlsx'
    eib_file_name = ''
    if not input_sheet:
        input_sheet, eib_file_name = load_file_params(load, base_path)

    # read Excel sheets
    #if load == "Absence Input":
        #input_data = pd.read_excel(input_data_temp_file, sheet_name=input_sheet, skiprows=[0, 1, 3, 4], dtype=object)
    #elif load == 'Worker Additional Data':
        #input_data = pd.read_excel(input_data_temp_file, sheet_name=input_sheet, skiprows=[0, 2, 3], dtype=object)
    #else:
    input_data = pd.read_excel(input_data_temp_file, sheet_name=input_sheet, skiprows=[1, 2], dtype=object)
    
    # initiate and laod the eib template that we need to write
    wb = load_workbook(eib_file_name)
    sheet_names = wb.sheetnames
    sheet_list = []

    # define the sheet number that we need to write the data in EIB template
    if load == "Change Job Assign Pay Group":
        sheet_name = wb[sheet_names[6]]
        sheet_list.append(sheet_name)
    
    elif load == "Change Job Change Org":
        sheet_name = wb[sheet_names[5]]
        sheet_list.append(sheet_name)
        input_data = input_data[
            ['Legacy Worker ID', 'Company Name', 'Business Area', 'Cost Center Code', 'Division', 'Current Business Area',
             'Current Cost Center Code', 'Current Division']]
        
    elif load in ("Payment Elections", "Absence Input", "Supervisory Org", "Prehire", "Put Candidate", "Supplier",
                  "Add Workday Account", "Company", "Update Workday Account","Job Classification", "Job Family", 
                  "Job Family Group", "Cost Center", "Location Hierarchy", "Custom Organizations", "Cost Center Hierarchy", "Collective Agreement",
                   "Put Supervisory Assignment Restrictions", "Location", "Job Category", "Job Profile", "Comp Grade and Grade Profile", "Skills",
                   "Job History Company", "Skills Reference Data", "Future Prehire"):
        sheet_name = wb[sheet_names[0]]
        sheet_list.append(sheet_name)
    
    # elif load == 'Job Requisition Roles':
    #     sheet_name = wb[sheet_names[3]]
    #     sheet_list.append(sheet_name)
    #     input_data = update_spreadsheet_key(eib_file_name, "Create Requisition", input_data)

    elif load in (
            'Edit Job Requisitions', 'Hire Employee', 'Performance Reviews', 'International Assignments',
            'Create Position','Hire CWR','Future Hire CWR', 'Overlapping Hire Employee', 'Future Hires', 'Change Job'):
        
        if load in ['Hire Employee', 'Overlapping Hire Employee', 'Future Hires']:
            hire_sheet = wb[sheet_names[1]]
            hire_assign_org_sheet = wb[sheet_names[8]]
            pay_group_sheet = wb[sheet_names[9]]
            service_date_sheet = wb[sheet_names[20]]
            sheet_list.append(hire_sheet)
            sheet_list.append(hire_assign_org_sheet)
            sheet_list.append(pay_group_sheet)
            sheet_list.append(service_date_sheet)
        
        elif load in ['Hire CWR', 'Future Hire CWR']:
            cwr_hire_sheet = wb[sheet_names[1]]
            cwr_hire_assign_org_sheet = wb[sheet_names[7]]
            sheet_list.append(cwr_hire_sheet)
            sheet_list.append(cwr_hire_assign_org_sheet)
            
        
        else:
            edit_job_sheet = wb[sheet_names[1]]
            assign_roles_sheet = wb[sheet_names[3]]
            sheet_list.append(edit_job_sheet)
            sheet_list.append(assign_roles_sheet)

            if load in ['Create Position', 'Change Job']:
                position_sheet = wb[sheet_names[2]]
                sheet_list.append(position_sheet)


    else:
        sheet_name = wb[sheet_names[1]]
        sheet_list.append(sheet_name)

    mapping_data_dict = get_mapping_data_dict(mapping_file)

    for sheet in sheet_list:
        # delete all the data from sheet from 6th row
        sheet.delete_rows(6, sheet.max_row)

        # method to convert our data and write it back to sheet
        sheet, all_unique_data_list, unavailable_reference_id_list = convert_data(input_data, sheet, load, mapping_file,
                                                                                  mapping_data_dict,
                                                                                  all_unique_data_list,
                                                                                  unavailable_reference_id_list,
                                                                                  input_data_temp_file, eib_file_name)
        # save the eib file and mark the load complete
        wb.save(eib_file_name)
        print("Completed " + load + " load")

    return all_unique_data_list, unavailable_reference_id_list

def get_eib_file_name(all_loads, base_path):
    eib_filename_list = []
    sheet_name_list = []
    for load in all_loads:
        input_sheet, eib_file_name = load_file_params(load, base_path)
        eib_filename_list.append(eib_file_name)
        sheet_name_list.append(input_sheet)
    return eib_filename_list, sheet_name_list

def creating_eib_files(all_loads, local_base_path, mapping_file, input_data_temp_file):
    
    global base_path
    base_path = local_base_path
    all_unique_data_list, unavailable_reference_id_list = pd.DataFrame(), pd.DataFrame()

    # hire, job req, request comp change, contingent worker hire, change job for contingent
    for load in all_loads:
        print("Starting load for: " + load)
        all_unique_data_list, unavailable_reference_id_list = process_load(load, all_unique_data_list,
                                                                           unavailable_reference_id_list, mapping_file,input_data_temp_file)

    # all_unique_data_list.dropna(subset=['Unique Value','Mapped Value'], inplace=True)
    if 'Unique Value' in all_unique_data_list.columns and 'Mapped Value' in all_unique_data_list.columns:
        all_unique_data_list.dropna(subset=['Unique Value','Mapped Value'], how='all', inplace=True)
    
    count_null_values = 0
    if 'Mapped Value' in all_unique_data_list.keys():
        count_null_values = all_unique_data_list['Mapped Value'].isnull().sum()

    if count_null_values > 0:
        print("Error: There are {} records where the value mapping is blank or null.".format(count_null_values))

    record_count = len(unavailable_reference_id_list)

    if record_count > 0:
        print("Error: The are {} mapping could not be found in the mapping file.".format(record_count))

    # #if count_null_values > 0 or record_count > 0:
    with pd.ExcelWriter('C:\\Source_Code\\unique_mapped_values.xlsx', engine='xlsxwriter') as writer:
        all_unique_data_list.to_excel(writer, sheet_name='unique_mapped_values', index=False)
        unavailable_reference_id_list.to_excel(writer, sheet_name='un_available_reference_type_id', index=False)


def creating_eib_files_v1(all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list):
    
    global base_path
    base_path = local_base_path
    all_unique_data_list, unavailable_reference_id_list = pd.DataFrame(), pd.DataFrame()

    # hire, job req, request comp change, contingent worker hire, change job for contingent
    for load in all_loads:
        print("Starting load for: " + load)
        all_unique_data_list, unavailable_reference_id_list = process_load(load, all_unique_data_list,
                                                                           unavailable_reference_id_list, mapping_file,input_data_temp_file)

    # all_unique_data_list.dropna(subset=['Unique Value','Mapped Value'], inplace=True)
    if 'Unique Value' in all_unique_data_list.columns and 'Mapped Value' in all_unique_data_list.columns:
        all_unique_data_list.dropna(subset=['Unique Value','Mapped Value'], how='all', inplace=True)
    
    count_null_values = 0
    if 'Mapped Value' in all_unique_data_list.keys():
        count_null_values = all_unique_data_list['Mapped Value'].isnull().sum()

    if count_null_values > 0:
        print("Error: There are {} records where the value mapping is blank or null.".format(count_null_values))

    record_count = len(unavailable_reference_id_list)

    if record_count > 0:
        print("Error: The are {} mapping could not be found in the mapping file.".format(record_count))

    return_all_unique_data_list.append(all_unique_data_list) 
    return_unavailable_reference_id_list.append(unavailable_reference_id_list)
    
    return return_all_unique_data_list, return_unavailable_reference_id_list

def main():
    # List of loads to process
    all_loads = ['Supervisory Org', 'Hire CWR', 'Prehire', 'Hire Employee', 'Job Requisitions',
                 'Change-Job', 'End Contingent Worker Contracts', 'Terminate Employee', 'Assign Work Schedule',
                 'Supplier', 'One Time Payments', 'Compensation', 'Probation Info', 'Edit Job Requisitions',
                 'Performance Reviews', 'Manage Goals', 'Leave Of Absence Events', 'Add Workday Account',
                 'Put Candidate', 'Role Based Assignments', 'Edit Position Restrictions Additional Data',
                 'NCA Additional Data Position', 'Worker Additional Data', 'International Assignments']


    #all_loads = ['Job Requisitions']
    all_loads = ['Hire CWR']
    #all_loads = ['Create Position', 'Hire Employee', 'Job Requisitions', 'Hire CWR', 'Prehire', 'Supervisory Org']

    # all_loads = ['Create Position', 'Hire Employee', 'Hire CWR', 'Prehire', 'Supervisory Org']
    all_unique_data_list, unavailable_reference_id_list = pd.DataFrame(), pd.DataFrame()
    # unavailable_reference_id_list = pd.DataFrame()

    global base_path
    # base_path = r"G:\\Shared drives\\NCS Australia Data Load\\"
    base_path = r"C:\\Source_Code\\"
    # hire, job req, request comp change, contingent worker hire, change job for contingent
    for load in all_loads:
        print("Starting load for: " + load)
        all_unique_data_list, unavailable_reference_id_list = process_load(load, all_unique_data_list,
                                                                           unavailable_reference_id_list)
    
    if "Mapped Value" in all_unique_data_list.keys():
        count_null_values = all_unique_data_list['Mapped Value'].isnull().sum()

    if count_null_values > 0:
        print("Error: There are {} records where the value mapping is blank or null.".format(count_null_values))

    record_count = len(unavailable_reference_id_list)

    if record_count > 0:
        print("Error: The are {} mapping could not be found in the mapping file.".format(record_count))

    #if count_null_values > 0 or record_count > 0:
    with pd.ExcelWriter(base_path + 'unique_mapped_values.xlsx', engine='xlsxwriter') as writer:
        all_unique_data_list.to_excel(writer, sheet_name='unique_mapped_values', index=False)
        unavailable_reference_id_list.to_excel(writer, sheet_name='un_available_reference_type_id', index=False)


# if __name__ == "__main__":
#     main()


import multiprocessing

def split_list(a_list):
    half = len(a_list)//2
    return a_list[:half], a_list[half:]


def creating_eib_files_with_parallel_processing(all_loads, local_base_path, mapping_file, input_data_temp_file):
    manager = multiprocessing.Manager()

    return_all_unique_data_list = manager.list()
    return_unavailable_reference_id_list = manager.list()
    # creating processes
    if len(all_loads) > 1:

        p1_all_loads, p2_all_loads = split_list(all_loads)
        print('>>>>>', p1_all_loads, p2_all_loads)
        
        p1 = multiprocessing.Process(target=creating_eib_files_v1, args=(p1_all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list,))
        p2 = multiprocessing.Process(target=creating_eib_files_v1, args=(p2_all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list,))
 
        # starting process 1
        p1.start()
        # starting process 2
        p2.start()

        # wait until process 1 is finished
        p1.join()
 
        # wait until process 2 is finished
        p2.join()

        print("Both processes finished Done!")
    else:
        creating_eib_files_v1(all_loads, local_base_path, mapping_file, input_data_temp_file, return_all_unique_data_list, return_unavailable_reference_id_list,)

    list_df_all_unique_data_list = [n for n in return_all_unique_data_list]
    
    list_df_unavailable_reference_id_list = [n for n in return_unavailable_reference_id_list]

    final_df_all_unique = pd.concat(list_df_all_unique_data_list, axis=0, ignore_index=True)
    final_df_unavailable_id = pd.concat(list_df_unavailable_reference_id_list, axis=0, ignore_index=True)
    
    with pd.ExcelWriter('C:\\Source_Code\\unique_mapped_values.xlsx', engine='xlsxwriter') as writer:
        final_df_all_unique.to_excel(writer, sheet_name='unique_mapped_values', index=False)
        final_df_unavailable_id.to_excel(writer, sheet_name='un_available_reference_type_id', index=False)

    print("EIB Files Creation Completed...")